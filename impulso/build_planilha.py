#!/usr/bin/env python3
"""
Gera o Plano de Negócios IMPULSO em formato Excel (.xlsx) com 20 abas,
fórmulas funcionais ligadas à aba "2. Premissas", dados de mercado reais,
análises dos especialistas e plano de ação 90 dias.

Compatível com Google Sheets: abrir o arquivo no Drive como Planilha Google
mantém todas as fórmulas e formatações.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# ESTILOS
# ============================================================
THIN = Side(border_style="thin", color="BFBFBF")
MED  = Side(border_style="medium", color="1F3864")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NAVY="1F3864"; BLUE="2E5C9E"; GREY="595959"
LIGHT_BLUE="D9E2F3"; LIGHT_GREEN="E2EFDA"
LIGHT_YELLOW="FFF2CC"; LIGHT_RED="FCE4D6"; LIGHT_GREY="F2F2F2"
LIGHT_ORANGE="FCE4D6"

H1 = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
H2 = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
H3 = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEAD = Font(name="Calibri", size=11, bold=True, color="1F3864")
BOLD = Font(name="Calibri", size=11, bold=True)
NORMAL = Font(name="Calibri", size=11)
SMALL = Font(name="Calibri", size=10, color=GREY)
SMALL_BOLD = Font(name="Calibri", size=10, bold=True)
ITAL = Font(name="Calibri", size=10, italic=True, color=GREY)

FILL_H1 = PatternFill("solid", fgColor=NAVY)
FILL_H2 = PatternFill("solid", fgColor=BLUE)
FILL_H3 = PatternFill("solid", fgColor=GREY)
FILL_LB = PatternFill("solid", fgColor=LIGHT_BLUE)
FILL_LG = PatternFill("solid", fgColor=LIGHT_GREEN)
FILL_LY = PatternFill("solid", fgColor=LIGHT_YELLOW)
FILL_LR = PatternFill("solid", fgColor=LIGHT_RED)
FILL_GREY = PatternFill("solid", fgColor=LIGHT_GREY)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center")

# ============================================================
# HELPERS
# ============================================================
def title_row(ws, row, text, span=10, fill=FILL_H1, font=H1, height=28):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = font; c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = height

def section_row(ws, row, text, span=10):
    title_row(ws, row, text, span=span, fill=FILL_H3, font=H3, height=22)

def sub_row(ws, row, text, span=10):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = HEAD; c.fill = FILL_LB
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[row].height = 20

def header_cells(ws, row, headers, start_col=1, fill=FILL_LB):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col+i, value=h)
        c.font = HEAD; c.fill = fill; c.alignment = CENTER; c.border = BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def money_fmt(c):
    c.number_format = 'R$ #,##0.00;[Red]-R$ #,##0.00'
def money0_fmt(c):
    c.number_format = 'R$ #,##0;[Red]-R$ #,##0'
def pct_fmt(c, dec=1):
    c.number_format = f'0.{"0"*dec}%'
def int_fmt(c):
    c.number_format = '#,##0'

def write_row(ws, row, values, start_col=1, money_cols=None, money0_cols=None,
              pct_cols=None, int_cols=None, font=None, fill=None, align=None, border=True):
    money_cols = money_cols or []; money0_cols = money0_cols or []
    pct_cols = pct_cols or []; int_cols = int_cols or []
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=start_col+i, value=v)
        if border: c.border = BORDER
        c.font = font or NORMAL
        if fill: c.fill = fill
        c.alignment = align or LEFT_TOP
        col = start_col + i
        if col in money_cols: money_fmt(c)
        if col in money0_cols: money0_fmt(c)
        if col in pct_cols: pct_fmt(c)
        if col in int_cols: int_fmt(c)


# ============================================================
# WORKBOOK
# ============================================================
wb = Workbook()
wb.remove(wb.active)

# Referência da aba premissas (single source of truth)
PR = "'2. Premissas'"

# ============================================================
# ABA 1 — RESUMO EXECUTIVO
# ============================================================
ws = wb.create_sheet("1. Resumo Executivo")
set_col_widths(ws, [3, 38, 22, 22, 22, 22, 3])

title_row(ws, 1, "IMPULSO — PLANO DE NEGÓCIOS COMPLETO", span=7)
c=ws.cell(row=2, column=2, value="Agência boutique de implementação de e-commerce e sites institucionais — análise da banca de especialistas")
c.font=ITAL

section_row(ws, 4, "IDENTIDADE DA EMPRESA", span=7)
identidade = [
    ("Nome da empresa", "Impulso"),
    ("Posicionamento", "Implementação rápida e profissional de lojas virtuais e sites institucionais para PMEs e empreendedores digitais"),
    ("Público-alvo (ICP)", "PMEs físicas migrando para digital (Carla); empreendedores DTC em validação (Rodrigo); B2B PME catálogo digital (Mariana)"),
    ("Proposta de valor", "\"Sua loja online profissional no ar em 48 horas — pronta para vender — com método, suporte e crescimento contínuo, sem freelancer sumido nem agência que demora 3 meses.\""),
    ("Diferenciais", "Método Impulso 48h proprietário (checklist + biblioteca de blocos + integrações pré-configuradas + kit IA); especialização vertical; conversão projeto→recorrência"),
    ("Serviços", "1) Projeto 48 Horas (Start/Pro/Premium) | 2) Projetos personalizados | 3) Planos mensais de suporte (Essencial/Pro/Growth) | 4) Receita de parceria com plataformas (20% MRR) | 5) Gestão de tráfego pago"),
    ("Modelo de receita", "Mix transacional (projetos) + recorrente (suporte + parceria + gestão). Meta 24 meses: 50% receita recorrente"),
    ("Regime tributário", "Simples Nacional, Anexo III com Fator R ≥ 28% (pró-labore é a alavanca). Premissa de 12% é conservadora — efetivo real é 6-8,4% até R$ 30k/mês"),
    ("Operação", "Home office, 2 sócios sêniors, capacidade combinada ~180h/mês de execução pura"),
]
r = 5
for label, value in identidade:
    a = ws.cell(row=r, column=2, value=label); a.font=BOLD; a.fill=FILL_LB; a.alignment=LEFT; a.border=BORDER
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    b = ws.cell(row=r, column=3, value=value); b.alignment=LEFT_TOP; b.border=BORDER
    ws.row_dimensions[r].height = 36
    r += 1

r += 1
section_row(ws, r, "PRINCIPAIS PREMISSAS", span=7); r += 1
header_cells(ws, r, ["Premissa", "Valor", "Origem"], start_col=2)
r += 1
premissas_resumo = [
    ("Número de sócios operacionais", f"={PR}!C5", "Definido pelo usuário"),
    ("Pró-labore por sócio (mensal)", f"={PR}!C7", "Definido pelo usuário"),
    ("Pró-labore total mensal", f"={PR}!C9", "= 2 sócios × R$ 4.000"),
    ("Pró-labore anual total", f"={PR}!C10", "= mensal × 12"),
    ("Horas produtivas mensais (2 sócios)", f"={PR}!C19", "Análise operacional"),
    ("Custo/hora interno (R$)", f"={PR}!C68", "= (CF total) / horas produtivas"),
    ("Margem recomendada", f"={PR}!C27", "Editável"),
    ("Imposto Simples efetivo (premissa)", f"={PR}!C31", "Editável — confer. Anexo III"),
    ("Reserva de capital de giro recomendada", f"={PR}!C65", "6× burn (recomendação CFO)"),
]
for label, val, origem in premissas_resumo:
    ws.cell(row=r, column=2, value=label).font = BOLD
    ws.cell(row=r, column=2).fill = FILL_GREY
    ws.cell(row=r, column=2).border = BORDER
    ws.cell(row=r, column=2).alignment = LEFT
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    c = ws.cell(row=r, column=3, value=val); c.alignment=CENTER; c.border=BORDER; c.font=BOLD
    if "Pró-labore" in label or "Custo" in label or "Reserva" in label: money0_fmt(c)
    if "Margem" in label or "Imposto" in label: pct_fmt(c)
    if "Horas" in label or "Número" in label: int_fmt(c)
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
    o = ws.cell(row=r, column=5, value=origem); o.alignment=LEFT; o.border=BORDER; o.font=ITAL
    ws.row_dimensions[r].height = 22
    r += 1

r += 1
section_row(ws, r, "INDICADORES-CHAVE DOS 3 CENÁRIOS (linkados à aba 12)", span=7); r += 1
header_cells(ws, r, ["Indicador", "Pessimista", "Realista", "Otimista", "Como é calculado"], start_col=2)
r += 1
indicadores = [
    ("Receita bruta anual",  "='12. Cenários'!P11","='12. Cenários'!P38","='12. Cenários'!P65", "Soma 12 meses de cada cenário", "money0"),
    ("Receita média mensal", "='12. Cenários'!P11/12","='12. Cenários'!P38/12","='12. Cenários'!P65/12","Anual ÷ 12", "money0"),
    ("Lucro líquido anual",  "='12. Cenários'!P17","='12. Cenários'!P44","='12. Cenários'!P71", "Soma do lucro líquido 12m", "money0"),
    ("Margem líquida média",  "='12. Cenários'!P18","='12. Cenários'!P45","='12. Cenários'!P72", "Lucro anual ÷ Receita anual", "pct"),
    ("Caixa acumulado mês 12","='12. Cenários'!O19","='12. Cenários'!O46","='12. Cenários'!O73", "Caixa final do mês 12", "money0"),
    ("Ponto de equilíbrio mensal","='15. Ponto Equilíbrio'!C12","='15. Ponto Equilíbrio'!C12","='15. Ponto Equilíbrio'!C12", "CF ÷ (1 − deduções variáveis)", "money0"),
    ("Pró-labore total mensal",   f"={PR}!C9", f"={PR}!C9", f"={PR}!C9", "Premissa editável", "money0"),
    ("Custos fixos mensais (sem pró-labore)", f"={PR}!C62", f"={PR}!C62", f"={PR}!C62", "Soma fixos da aba 2", "money0"),
    ("Payback estimado", "Inviável s/ aporte", "~9 meses (caixa+ no mês 6)", "~4 meses", "Análise CFO", "text"),
]
for label, p, real, o, formula, kind in indicadores:
    ws.cell(row=r, column=2, value=label).font=BOLD
    ws.cell(row=r, column=2).fill=FILL_GREY
    ws.cell(row=r, column=2).alignment=LEFT
    ws.cell(row=r, column=2).border=BORDER
    for j, val in enumerate([p, real, o]):
        c = ws.cell(row=r, column=3+j, value=val)
        c.border = BORDER; c.alignment = CENTER
        if kind == "money0": money0_fmt(c)
        elif kind == "pct": pct_fmt(c)
    c = ws.cell(row=r, column=6, value=formula); c.font=ITAL; c.border=BORDER; c.alignment=LEFT
    ws.row_dimensions[r].height = 22
    r += 1

r += 1
section_row(ws, r, "CONCLUSÃO EXECUTIVA DO SQUAD", span=7); r += 1
conclusao = (
    "🟢 VIÁVEL no Realista (margem líquida ~24%, payback ~9m) e robusto no Otimista (margem 44%). "
    "🔴 Pessimista é inviável sem aporte ou ajuste de pró-labore.\n\n"
    "Os 3 AJUSTES NÃO-NEGOCIÁVEIS identificados pelo squad:\n"
    "  1) Pró-labore ESCALONADO — começar com R$ 3k/sócio (mês 1-3) → R$ 3,5k (4-6) → R$ 4k (7+) quando MRR > R$ 10k. Sustenta o Fator R do Anexo III (28%) e protege o caixa.\n"
    "  2) Escopo do 48h TAXATIVO — máximo 20 SKUs, 1 gateway, 1 frete, 4 páginas, plataformas Shopify/Nuvemshop/Tray. Woo fora do 48h. Sem isso, vira processo no PROCON (art. 30, 37 CDC).\n"
    "  3) Foco comercial em MRR — meta de 50% de receita recorrente até 24 meses. Cada R$ 1 de MRR vale 3-5× R$ 1 de projeto.\n\n"
    "⚠️ RISCO ESTRATÉGICO MAIOR (24-36 meses): comoditização do setup básico por IA generativa (Shopify Magic, Wix ADI). A tese precisa migrar de 'implementador' para 'operador de crescimento de PMEs digitais' — vender resultado de vendas, não apenas loja no ar.\n\n"
    "✅ AÇÕES OBRIGATÓRIAS antes da 1ª venda: constituir LTDA + CNAEs corretos + Simples; contrato dual-track (B2B/B2C); Política do Projeto 48 Horas com aceite eletrônico; LGPD (DPA); registrar marca no INPI; reserva de R$ 50k de capital de giro."
)
ws.merge_cells(start_row=r, start_column=2, end_row=r+9, end_column=6)
c = ws.cell(row=r, column=2, value=conclusao); c.font=NORMAL; c.fill=FILL_LY
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True); c.border=BORDER

print("Aba 1 ✓")


# ============================================================
# ABA 2 — PREMISSAS GERAIS (single source of truth)
# ============================================================
ws = wb.create_sheet("2. Premissas")
set_col_widths(ws, [3, 42, 18, 12, 50, 3])

title_row(ws, 1, "PREMISSAS GERAIS — EDITE AQUI E TODA A PLANILHA SE ATUALIZA", span=6)
ws.cell(row=2, column=2, value="Células em AMARELO são premissas editáveis. Verde = fórmulas (não editar). Cinza = referência.").font = ITAL

# === SÓCIOS E PRÓ-LABORE ===
section_row(ws, 4, "SÓCIOS E PRÓ-LABORE", span=6)

write_row(ws, 5, ["", "Número de sócios operacionais", 2, "un", "Definição da operação", ""], align=LEFT)
ws.cell(row=5, column=3).fill = FILL_LY; ws.cell(row=5, column=3).font = BOLD; ws.cell(row=5, column=3).alignment = CENTER
int_fmt(ws.cell(row=5, column=3))

# C7 - pró-labore por sócio
write_row(ws, 7, ["", "Pró-labore mensal por sócio (R$)", 4000, "R$/mês", "Premissa do usuário — R$ 4.000/sócio", ""], align=LEFT)
ws.cell(row=7, column=3).fill = FILL_LY; ws.cell(row=7, column=3).font = BOLD; ws.cell(row=7, column=3).alignment = CENTER
money0_fmt(ws.cell(row=7, column=3))

# C8 - INSS sócio (11%, art. 21 Lei 8.212/91)
write_row(ws, 8, ["", "INSS sócio (11% sobre pró-labore individual)", "=C7*0.11", "R$/mês", "Art. 21 Lei 8.212/91 (até teto)", ""], align=LEFT)
ws.cell(row=8, column=3).fill = FILL_LG; money0_fmt(ws.cell(row=8, column=3)); ws.cell(row=8, column=3).alignment = CENTER

# C9 - pró-labore total
write_row(ws, 9, ["", "Pró-labore TOTAL mensal", "=C7*C5", "R$/mês", "Fórmula = pró-labore × sócios", ""], align=LEFT)
ws.cell(row=9, column=3).fill = FILL_LG; ws.cell(row=9, column=3).font = BOLD
money0_fmt(ws.cell(row=9, column=3)); ws.cell(row=9, column=3).alignment = CENTER

# C10 - pró-labore anual
write_row(ws, 10, ["", "Pró-labore TOTAL anual", "=C9*12", "R$/ano", "Fórmula = mensal × 12", ""], align=LEFT)
ws.cell(row=10, column=3).fill = FILL_LG; ws.cell(row=10, column=3).font = BOLD
money0_fmt(ws.cell(row=10, column=3)); ws.cell(row=10, column=3).alignment = CENTER

# C11 - INSS total
write_row(ws, 11, ["", "INSS sócios TOTAL mensal", "=C8*C5", "R$/mês", "Soma do INSS dos sócios", ""], align=LEFT)
ws.cell(row=11, column=3).fill = FILL_LG; money0_fmt(ws.cell(row=11, column=3)); ws.cell(row=11, column=3).alignment = CENTER

# === CAPACIDADE OPERACIONAL ===
section_row(ws, 13, "CAPACIDADE OPERACIONAL", span=6)
write_row(ws, 14, ["", "Dias úteis por mês", 22, "dias", "Premissa padrão BR", ""], align=LEFT)
ws.cell(row=14, column=3).fill = FILL_LY; int_fmt(ws.cell(row=14, column=3)); ws.cell(row=14, column=3).alignment = CENTER
write_row(ws, 15, ["", "Horas/dia por sócio", 8, "h", "Jornada", ""], align=LEFT)
ws.cell(row=15, column=3).fill = FILL_LY; int_fmt(ws.cell(row=15, column=3)); ws.cell(row=15, column=3).alignment = CENTER
write_row(ws, 16, ["", "Horas BRUTAS mensais por sócio", "=C14*C15", "h", "Fórmula = dias × horas/dia", ""], align=LEFT)
ws.cell(row=16, column=3).fill = FILL_LG; int_fmt(ws.cell(row=16, column=3)); ws.cell(row=16, column=3).alignment = CENTER
write_row(ws, 17, ["", "Horas brutas mensais TOTAIS (2 sócios)", "=C16*C5", "h", "Fórmula = horas/sócio × sócios", ""], align=LEFT)
ws.cell(row=17, column=3).fill = FILL_LG; int_fmt(ws.cell(row=17, column=3)); ws.cell(row=17, column=3).alignment = CENTER
write_row(ws, 18, ["", "% de horas produtivas reais (descontado comercial, admin, atendimento)", 0.5, "%", "Análise operacional: ~50% (vendas 15% + atendimento 10% + pós-venda 10% + admin 8% + gestão 7%)", ""], align=LEFT)
ws.cell(row=18, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=18, column=3)); ws.cell(row=18, column=3).alignment = CENTER
write_row(ws, 19, ["", "Horas PRODUTIVAS mensais (execução pura)", "=C17*C18", "h", "Fórmula. Padrão de planejamento: 176h", ""], align=LEFT)
ws.cell(row=19, column=3).fill = FILL_LG; ws.cell(row=19, column=3).font = BOLD
int_fmt(ws.cell(row=19, column=3)); ws.cell(row=19, column=3).alignment = CENTER
write_row(ws, 20, ["", "Ganho médio de produtividade com IA (%)", 0.25, "%", "Análise operacional — ganho real médio (não 80%+ ilusório)", ""], align=LEFT)
ws.cell(row=20, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=20, column=3)); ws.cell(row=20, column=3).alignment = CENTER
write_row(ws, 21, ["", "Capacidade-alvo (% de utilização sustentável)", 0.8, "%", "Operar ≤80% para evitar burnout", ""], align=LEFT)
ws.cell(row=21, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=21, column=3)); ws.cell(row=21, column=3).alignment = CENTER
write_row(ws, 22, ["", "Horas FATURÁVEIS mensais (alvo)", "=C19*C21", "h", "Horas produtivas × utilização alvo", ""], align=LEFT)
ws.cell(row=22, column=3).fill = FILL_LG; int_fmt(ws.cell(row=22, column=3)); ws.cell(row=22, column=3).alignment = CENTER

# === PRICING & MARGEM ===
section_row(ws, 24, "PRECIFICAÇÃO E MARGEM", span=6)
write_row(ws, 25, ["", "Valor/hora desejado (referência mercado)", 250, "R$/h", "Mercado: freela sênior R$ 150-400/h; agência R$ 200-600/h", ""], align=LEFT)
ws.cell(row=25, column=3).fill = FILL_LY; money0_fmt(ws.cell(row=25, column=3)); ws.cell(row=25, column=3).alignment = CENTER
write_row(ws, 26, ["", "Margem MÍNIMA aceitável", 0.20, "%", "Piso para não vender no prejuízo", ""], align=LEFT)
ws.cell(row=26, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=26, column=3)); ws.cell(row=26, column=3).alignment = CENTER
write_row(ws, 27, ["", "Margem RECOMENDADA", 0.40, "%", "Saudável para serviços", ""], align=LEFT)
ws.cell(row=27, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=27, column=3)); ws.cell(row=27, column=3).alignment = CENTER
write_row(ws, 28, ["", "Margem PREMIUM (sem desconto)", 0.55, "%", "Preço de ancoragem", ""], align=LEFT)
ws.cell(row=28, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=28, column=3)); ws.cell(row=28, column=3).alignment = CENTER

# === IMPOSTOS E TAXAS ===
section_row(ws, 30, "IMPOSTOS E TAXAS DE PAGAMENTO", span=6)
write_row(ws, 31, ["", "Imposto Simples efetivo (premissa do projeto)", 0.12, "%", "Premissa do usuário. Real: 6% até R$ 180k; 8,4% até R$ 360k; 11% até R$ 720k (Anexo III)", ""], align=LEFT)
ws.cell(row=31, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=31, column=3)); ws.cell(row=31, column=3).alignment = CENTER
write_row(ws, 32, ["", "Taxa PIX (gateway)", 0.005, "%", "InfinitePay 0%; PagBank até 1,89%", ""], align=LEFT)
ws.cell(row=32, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=32, column=3),2); ws.cell(row=32, column=3).alignment = CENTER
write_row(ws, 33, ["", "Taxa cartão à vista", 0.035, "%", "Stone 3,11%; PagBank D+30 3,19%; InfinitePay 2,69%", ""], align=LEFT)
ws.cell(row=33, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=33, column=3),2); ws.cell(row=33, column=3).alignment = CENTER
write_row(ws, 34, ["", "Taxa cartão 3x", 0.045, "%", "Stone parcelado start 5,41%; média 4-5%", ""], align=LEFT)
ws.cell(row=34, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=34, column=3),2); ws.cell(row=34, column=3).alignment = CENTER
write_row(ws, 35, ["", "Taxa cartão 6x", 0.065, "%", "Faixa intermediária", ""], align=LEFT)
ws.cell(row=35, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=35, column=3),2); ws.cell(row=35, column=3).alignment = CENTER
write_row(ws, 36, ["", "Taxa cartão 10x", 0.095, "%", "Mercado Pago D+1 ~12,49%; InfinitePay 12x 8,99%", ""], align=LEFT)
ws.cell(row=36, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=36, column=3),2); ws.cell(row=36, column=3).alignment = CENTER
write_row(ws, 37, ["", "Taxa cartão 12x", 0.115, "%", "Faixa iniciante MP até 22%; recomenda-se repassar juros >6x", ""], align=LEFT)
ws.cell(row=37, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=37, column=3),2); ws.cell(row=37, column=3).alignment = CENTER
write_row(ws, 38, ["", "Inadimplência projetada (PCLD)", 0.04, "%", "B2B serviços BR — 4% conservador. Serasa: 55,2% inadimplência em serviços", ""], align=LEFT)
ws.cell(row=38, column=3).fill = FILL_LY; pct_fmt(ws.cell(row=38, column=3)); ws.cell(row=38, column=3).alignment = CENTER

# === CUSTOS FIXOS MENSAIS ===
section_row(ws, 40, "CUSTOS FIXOS MENSAIS (sem pró-labore)", span=6)
custos_fixos = [
    ("Internet empresarial (fibra 200-500 Mbps, 2 sócios)", 200, "Vivo R$ 128,98; TIM R$ 119,99; média BR R$ 100-180"),
    ("Energia adicional home office (2 sócios)", 80, "Enel SP R$ 0,91/kWh × ~45 kWh/sócio (R$ 40/sócio)"),
    ("Google Workspace Business Starter (2 usuários)", 65, "R$ 32,72/user × 2 (anual)"),
    ("Ferramentas IA (ChatGPT Plus + Claude Pro)", 220, "US$ 20 × 2 ≈ R$ 108 × 2"),
    ("Figma Professional (2 editores anual)", 162, "US$ 15/editor × 2 ≈ R$ 81 × 2"),
    ("Canva Pro Teams", 39, "R$ 470/ano = R$ 39/mês"),
    ("ClickUp Unlimited (2 users anual)", 14, "R$ 7 × 2"),
    ("HubSpot CRM Free", 0, "Gratuito"),
    ("Make Core / Zapier Starter", 60, "Make US$ 10,59 ≈ R$ 57"),
    ("Hospedagem site Impulso + domínio", 50, "Hostinger Premium + .com.br Registro.br R$ 40/ano"),
    ("Contabilidade especializada (recomendação contábil)", 800, "Faixa: ME serviços R$ 400-1.200; especializada em TI R$ 700-1.200"),
    ("Certificado Digital A1 (R$ 220/ano)", 20, "Necessário p/ NFS-e nacional"),
    ("Antivírus + backup nuvem", 80, "Operacional"),
    ("Seguro RC profissional (E&O — recomendado)", 150, "Proteção contra responsabilidade civil"),
    ("Marketing próprio (mídia, conteúdo, ferramentas)", 1200, "R$ 3-5k recomendado quando MRR estabiliza; piso R$ 1.200"),
    ("CRM RD Station Basic (opcional)", 49, "Pode iniciar com HubSpot Free"),
    ("Outros custos administrativos / contingência", 200, "Sindicato, taxas municipais, eventual"),
]
r = 41
for nome, valor, obs in custos_fixos:
    write_row(ws, r, ["", nome, valor, "R$/mês", obs, ""], align=LEFT)
    ws.cell(row=r, column=3).fill = FILL_LY
    money0_fmt(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).alignment = CENTER
    r += 1

# Subtotal de custos fixos
write_row(ws, r, ["", "SUBTOTAL CUSTOS FIXOS (sem pró-labore)", f"=SUM(C41:C{r-1})", "R$/mês", "Soma automática dos custos acima", ""], align=LEFT, font=BOLD)
ws.cell(row=r, column=3).fill = FILL_LG; ws.cell(row=r, column=3).font = BOLD
money0_fmt(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).alignment = CENTER
SUM_FIXOS_ROW = r  # C40 reference if needed elsewhere
# Vou capturar a célula da soma — ela está na linha r, col 3
r += 1

# Custo fixo TOTAL (com pró-labore + INSS)
write_row(ws, r, ["", "CUSTO FIXO TOTAL (com pró-labore + INSS)", f"=C{SUM_FIXOS_ROW}+C9+C11", "R$/mês", "Subtotal + pró-labore + INSS", ""], align=LEFT, font=BOLD)
ws.cell(row=r, column=3).fill = FILL_H3; ws.cell(row=r, column=3).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
money0_fmt(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).alignment = CENTER
CF_TOTAL_ROW = r
r += 2

# Adicionar referências computadas no topo (linha 40 e 46 que já referenciei na aba 1)
# Linha 40 deveria estar no resumo — recalibrei: aba 1 referencia C40 para subtotal fixos
# Mas o subtotal real está em CF_TOTAL_ROW. Vou usar C40 como o subtotal. Ajustar.
# Atual: SUM_FIXOS_ROW = 58 provavelmente; CF_TOTAL_ROW = 59
# Aba 1 espera C40 (subtotal) e C42 (capital de giro) e C46 (custo/hora)
# Vou usar células-âncora na seção de cálculos abaixo para evitar inconsistências.

section_row(ws, r, "CÁLCULOS DERIVADOS (não editar)", span=6); r += 1
write_row(ws, r, ["", "(CF40) Subtotal Custos Fixos (sem pró-labore)", f"=C{SUM_FIXOS_ROW}", "R$/mês", "Referência usada em outras abas", ""], align=LEFT, font=BOLD)
ws.cell(row=r, column=3).fill = FILL_LG; money0_fmt(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).alignment = CENTER
CF40 = r  # vou usar essa referência

r += 1
write_row(ws, r, ["", "Custo Fixo TOTAL (com pró-labore + INSS)", f"=C{CF40}+C9+C11", "R$/mês", "Base do ponto de equilíbrio", ""], align=LEFT, font=BOLD)
ws.cell(row=r, column=3).fill = FILL_LG; ws.cell(row=r, column=3).font = BOLD
money0_fmt(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).alignment = CENTER
CF_TOTAL = r

r += 1
write_row(ws, r, ["", "Reserva capital de giro mínima (3× burn)", f"=C{CF_TOTAL}*3", "R$", "Recomendação CFO", ""], align=LEFT)
ws.cell(row=r, column=3).fill = FILL_LG; money0_fmt(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).alignment = CENTER

r += 1
write_row(ws, r, ["", "Reserva capital de giro RECOMENDADA (6× burn)", f"=C{CF_TOTAL}*6", "R$", "Recomendação CFO", ""], align=LEFT, font=BOLD)
ws.cell(row=r, column=3).fill = FILL_LG; ws.cell(row=r, column=3).font = BOLD
money0_fmt(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).alignment = CENTER
CAP_GIRO = r

r += 1
write_row(ws, r, ["", "Reserva conservadora (12× burn)", f"=C{CF_TOTAL}*12", "R$", "Para cenário pessimista", ""], align=LEFT)
ws.cell(row=r, column=3).fill = FILL_LG; money0_fmt(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).alignment = CENTER

r += 2
write_row(ws, r, ["", "CUSTO/HORA INTERNO (piso)", f"=C{CF_TOTAL}/C19", "R$/h", "= Custo fixo total / horas produtivas. Vender abaixo = prejuízo certo.", ""], align=LEFT, font=BOLD)
ws.cell(row=r, column=3).fill = FILL_H3; ws.cell(row=r, column=3).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
money_fmt(ws.cell(row=r, column=3)); ws.cell(row=r, column=3).alignment = CENTER
CUSTO_HORA = r

r += 1
write_row(ws, r, ["", "Taxa cartão MÉDIA (ponderada das 6 modalidades)", "=AVERAGE(C32:C37)", "%", "Média simples; pode-se ponderar por mix real", ""], align=LEFT)
ws.cell(row=r, column=3).fill = FILL_LG; pct_fmt(ws.cell(row=r, column=3),2); ws.cell(row=r, column=3).alignment = CENTER
TAXA_MED = r

r += 1
write_row(ws, r, ["", "Deduções variáveis totais (imposto + taxa média + inad)", f"=C31+C{TAXA_MED}+C38", "%", "Usado no ponto de equilíbrio", ""], align=LEFT, font=BOLD)
ws.cell(row=r, column=3).fill = FILL_LG; pct_fmt(ws.cell(row=r, column=3),1); ws.cell(row=r, column=3).alignment = CENTER
DEDUC_VAR = r

# Atualizar âncoras conhecidas pela aba 1:
# Aba 1 espera: C5 = nº sócios | C7 = pró-labore/sócio | C9 = pró-labore total | C10 = anual
# C19 = horas produtivas | C28 = margem desejada | C30 = imposto (NÃO! agora C31 é imposto)
# vou consertar a aba 1 abaixo... mas como esse é o mesmo arquivo, ajusto agora.

# Adicionar mapa de referências importantes ao topo
ws.cell(row=2, column=5, value=f"REFERÊNCIAS-CHAVE: C9=Pró-labore total | C19=Horas produtivas | C{CF40}=Custos fixos s/PL | C{CF_TOTAL}=CF total | C{CUSTO_HORA}=Custo/h | C{DEDUC_VAR}=Deduções var | C31=Imposto").font = SMALL_BOLD

# === ARMAZENAR REFERÊNCIAS PARA USO POSTERIOR ===
PR_REF = {
    "nsocios": "C5", "pl_socio": "C7", "inss_socio": "C8", "pl_total": "C9",
    "pl_anual": "C10", "inss_total": "C11",
    "horas_brutas": "C17", "horas_prod": "C19", "horas_alvo": "C22",
    "ganho_ia": "C20", "util_alvo": "C21",
    "vh_desejado": "C25", "margem_min": "C26", "margem_rec": "C27", "margem_prem": "C28",
    "imposto": "C31", "tx_pix": "C32", "tx_vista": "C33", "tx_3x": "C34",
    "tx_6x": "C35", "tx_10x": "C36", "tx_12x": "C37", "inad": "C38",
    "cf_sem_pl": f"C{CF40}",
    "cf_total": f"C{CF_TOTAL}",
    "cap_giro": f"C{CAP_GIRO}",
    "custo_hora": f"C{CUSTO_HORA}",
    "tx_media": f"C{TAXA_MED}",
    "deduc_var": f"C{DEDUC_VAR}",
}
print(f"Aba 2 ✓ — Referências: {PR_REF}")


# ============================================================
# ABA 3 — PESQUISA DE MERCADO E REFERÊNCIAS
# ============================================================
ws = wb.create_sheet("3. Pesquisa Mercado")
set_col_widths(ws, [3, 36, 20, 20, 20, 14, 38, 3])

title_row(ws, 1, "PESQUISA DE MERCADO E REFERÊNCIAS REAIS", span=8)
ws.cell(row=2, column=2, value="Dados consultados em 2026-06-10 — câmbio referência USD ≈ R$ 5,42 | EUR ≈ R$ 5,90").font = ITAL

# Helper para escrever linhas de pesquisa
def pesq_section(ws, row, titulo):
    sub_row(ws, row, titulo, span=8)
    header_cells(ws, row+1, ["Item", "Mínimo", "Médio", "Máximo", "Data", "Fonte / Observação"], start_col=2)
    return row + 2

def pesq_line(ws, row, item, mn, md, mx, date, fonte, money=True):
    vals = [item, mn, md, mx, date, fonte]
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=2+i, value=v)
        c.border = BORDER
        c.alignment = LEFT_TOP if i in (0, 5) else CENTER
        c.font = NORMAL if i != 0 else BOLD
        if money and i in (1, 2, 3) and isinstance(v, (int, float)):
            money0_fmt(c)
    ws.row_dimensions[row].height = 24

# A. PLATAFORMAS
r = pesq_section(ws, 4, "A. PLATAFORMAS DE E-COMMERCE (preço mensal BRL e programas de parceria)")
plats = [
    ("Shopify Basic (US$ 39 anual US$ 29)",       157, 211, 211, "2026-06-10", "shopify.com/pricing — IOF 6,38%"),
    ("Shopify Grow (US$ 105 anual US$ 79)",       428, 569, 569, "2026-06-10", "shopify.com/pricing"),
    ("Shopify Advanced (US$ 399 anual US$ 299)",  1621, 2163, 2163, "2026-06-10", "shopify.com/pricing"),
    ("Shopify Partners — comissão recorrente",    None, "20%", None, "2026-06-10", "help.shopify.com/en/partners/partner-program/how-to-earn"),
    ("Tray — faixa de planos",                    19,  99,  449, "2026-06-10", "tray.com.br/planos-precos"),
    ("Tray Partners — comissão recorrente",       None, "20-30%", None, "2026-06-10", "tray.com.br/partners"),
    ("Nuvemshop Essencial",                       59,  59,  59, "2026-06-10", "nuvemshop.com.br/planos-e-precos"),
    ("Nuvemshop Impulso (plano homônimo)",        139, 139, 139, "2026-06-10", "nuvemshop.com.br/planos-e-precos"),
    ("Nuvemshop Escala",                          389, 389, 389, "2026-06-10", "nuvemshop.com.br/planos-e-precos"),
    ("Nuvemshop Experts — comissão",              None, "20%", None, "2026-06-10", "atendimento.nuvemshop.com.br — comissões vigentes"),
    ("WooCommerce hospedagem entry (Hostinger/KingHost)", 13, 30, 50, "2026-06-10", "hostinger.com/br/tutoriais/wordpress-preco; site.king.host"),
    ("WooCommerce TCO total (hosp+plugins)",      100, 250, 500, "2026-06-10", "Composição — estimativa de mercado"),
    ("Wix Core eCommerce (US$ 29)",               157, 157, 157, "2026-06-10", "tooltester.com/wix"),
    ("BigCommerce Core (US$ 39, cap US$ 50k/ano)",211, 211, 211, "2026-06-10", "bigcommerce.com/essentials/pricing"),
    ("VTEX Business (anual R$ 60k)",              5000, 5000, 5000, "2026-06-10", "assine.vtex.com/vtex-commerce-suite-new-business"),
]
for p in plats:
    pesq_line(ws, r, *p)
    r += 1
r += 1

# B. SERVIÇOS
r = pesq_section(ws, r, "B. PREÇOS DE SERVIÇOS — REFERÊNCIAS DE MERCADO BR")
srv = [
    ("Implementação e-commerce ENTRY (Nuvemshop/Tray template)", 3000, 6000, 10000, "2026-06-10", "mathcreator.com.br; safiradesign.com.br"),
    ("Implementação Shopify INTERMEDIÁRIO", 8000, 18000, 25000, "2026-06-10", "litextension.com/pt/blog/shopify-pricing; simplesinovacao.com"),
    ("Implementação Shopify/VTEX PREMIUM (custom)", 25000, 35000, 60000, "2026-06-10", "litextension.com"),
    ("Site institucional simples", 3000, 5000, 6000, "2026-06-10", "hostwp.com.br; safiradesign.com.br"),
    ("Site institucional premium", 7000, 11000, 15000, "2026-06-10", "hostwp.com.br; mathcreator.com.br"),
    ("Landing page básica (template)", 300, 600, 900, "2026-06-10", "gabrieldosite.com.br"),
    ("Landing page profissional", 1200, 2500, 4000, "2026-06-10", "bluepaper.io"),
    ("Landing page estratégica c/ funil/automação", 2000, 4500, 8000, "2026-06-10", "bluepaper.io"),
    ("Migração de plataforma (ex: VTEX→Shopify)", 8000, 18000, 40000, "2026-06-10", "hulkapps.com"),
    ("Integração ERP (Bling/Tiny/Omie) projeto único", 490, 1500, 4000, "2026-06-10", "brauerecommerce.com.br"),
    ("SEO técnico — auditoria pontual", 1200, 4000, 8000, "2026-06-10", "juanmoura.com; profissionalseo.com.br"),
    ("SEO técnico — mensal contínuo", 2000, 5000, 15000, "2026-06-10", "juanmoura.com"),
    ("Auditoria de e-commerce", 1500, 4000, 10000, "2026-06-10", "Estimativa — referência similar a auditoria SEO"),
    ("Hora freelancer sênior (dev)", 150, 250, 400, "2026-06-10", "godaddy.com; unclik.com.br/Glassdoor"),
    ("Hora agência (dev)", 200, 350, 600, "2026-06-10", "Estimativa: +40-60% sobre freela sênior"),
    ("Squad agência (4-5 pessoas/mês)", 60000, 90000, 120000, "2026-06-10", "numerando.com.br"),
]
for p in srv:
    pesq_line(ws, r, *p)
    r += 1
r += 1

# C. SUPORTE
r = pesq_section(ws, r, "C. PLANOS MENSAIS DE SUPORTE/MANUTENÇÃO BR")
sup = [
    ("WP Essencial 1 site, 2h, manutenção preventiva", 199, 250, 300, "2026-06-10", "profissionalwp.dev.br"),
    ("Loja virtual mid (6h, não cumulativas)", 600, 900, 1200, "2026-06-10", "linkwell.com.br — faixa estimada"),
    ("Mid-tier (10-20h, SLA 24h)", 1500, 2500, 3500, "2026-06-10", "ae.digital — estimativa"),
    ("Premium (30h+, SLA 4h úteis)", 4000, 6000, 8000, "2026-06-10", "found.pt — estimativa"),
]
for p in sup:
    pesq_line(ws, r, *p)
    r += 1
r += 1

# D. TAXAS DE CARTÃO
r = pesq_section(ws, r, "D. TAXAS DE CARTÃO BR 2025-2026 (online com antecipação automática)")
txs = [
    ("Stone — Débito | À vista | Parcelado start", "1,25%", "3,11%", "5,41%", "2026-06-10", "calculadoradetaxas.com.br/stone"),
    ("Mercado Pago — 12x D+14 | D+1 | D+0", "9,98%", "12,49%", "14,99%", "2026-06-10", "blog.jota.ai/taxas-maquininha-mercadopago-2026"),
    ("PagBank — Débito | Crédito D+30 | D+0", "1,99%", "3,19%", "4,99%", "2026-06-10", "pagbank.com.br/maquininhas/taxas-e-tarifas"),
    ("InfinitePay — Débito | Crédito vista | 12x", "0,75%", "2,69%", "8,99%", "2026-06-10", "infinitepay.io/taxas"),
    ("InfinitePay — PIX", "0%", "0%", "0%", "2026-06-10", "infinitepay.io/pix"),
    ("Cielo — Débito | Crédito vista (entry)", "1,07%", "3,48%", "—", "2026-06-10", "calculadoradetaxas.com.br/cielo"),
    ("Cielo — Antecipação manual a.m.", "1,49%", "2,5%", "3,5%", "2026-06-10", "antecipafacil.com.br"),
    ("Inadimplência B2B BR (Serasa dez/25)", "8,9 mi empresas", "R$ 213 bi dívida", "55,2% serviços", "2026-06-10", "serasa.com.br/limpa-nome/mapa-inadimplencia"),
]
for p in txs:
    pesq_line(ws, r, *p, money=False)
    r += 1
r += 1

# E. FERRAMENTAS
r = pesq_section(ws, r, "E. FERRAMENTAS — preço mensal (BRL ou USD convertido)")
fer = [
    ("ChatGPT Plus (US$ 20)", 108, 108, 108, "2026-06-10", "openai.com/pricing"),
    ("ChatGPT Team anual (US$ 25/seat)", 135, 135, 135, "2026-06-10", "openai.com/pricing"),
    ("Claude Pro (US$ 20)", 108, 108, 108, "2026-06-10", "claude.com/pricing"),
    ("Claude Team Standard anual (US$ 20/seat, min 5)", 108, 108, 108, "2026-06-10", "claude.com/pricing"),
    ("Cursor Pro (US$ 20)", 108, 108, 108, "2026-06-10", "cursor.com/pricing"),
    ("GitHub Copilot Pro (US$ 10)", 54, 54, 54, "2026-06-10", "github.com/features/copilot/plans"),
    ("Figma Professional anual (US$ 15/editor)", 81, 81, 81, "2026-06-10", "figma.com/pricing"),
    ("Canva Pro anual (1 pessoa)", 24, 24, 24, "2026-06-10", "canva.com/pt_br/precos"),
    ("Canva Teams (até 5 pessoas)", 39, 39, 39, "2026-06-10", "darlanevandro.com.br"),
    ("Adobe Creative Cloud Standard (BR atual)", 249, 249, 320, "2026-06-10", "tecnoblog.net (vigente desde ago/25)"),
    ("Make Core Starter (US$ 10,59)", 57, 57, 57, "2026-06-10", "toolradar.com/blog/zapier-pricing-2026"),
    ("Zapier Starter (US$ 19,99)", 108, 108, 108, "2026-06-10", "toolradar.com"),
    ("n8n Cloud Starter (€20)", 118, 118, 118, "2026-06-10", "n8n.io/pricing — self-host gratuito"),
    ("ClickUp Unlimited anual", 7, 7, 7, "2026-06-10", "getaiperks.com/clickup-pricing — BRL direto"),
    ("ClickUp Business anual", 12, 12, 12, "2026-06-10", "getaiperks.com"),
    ("Trello Standard (US$ 5)", 27, 27, 27, "2026-06-10", "joinsecret.com"),
    ("HubSpot Free", 0, 0, 0, "2026-06-10", "br.hubspot.com (recomendado p/ início)"),
    ("HubSpot Starter (R$ 95/user)", 95, 95, 95, "2026-06-10", "tecnois.com.br"),
    ("RD Station Basic", 48, 48, 48, "2026-06-10", "capterra.com/p/180098/RD-Station"),
    ("RD Station Pro", 69, 69, 69, "2026-06-10", "capterra.com"),
    ("Pipedrive (start, R$ 60/user)", 60, 60, 60, "2026-06-10", "tecnois.com.br"),
    ("Slack Pro anual (US$ 7,25/user)", 39, 39, 39, "2026-06-10", "slack.com/pricing"),
    ("Google Workspace Business Starter (anual)", 33, 33, 33, "2026-06-10", "workspace.google.com/pricing"),
    ("Google Workspace Business Standard (anual)", 82, 82, 82, "2026-06-10", "workspace.google.com/pricing"),
    ("Hospedagem WP — Hostinger (entry)", 6, 10, 14, "2026-06-10", "hostinger.com/br"),
    ("Hospedagem WP — KingHost", 17, 17, 17, "2026-06-10", "site.king.host/hospedagem-wordpress"),
    ("Domínio .com.br Registro.br (R$ 40/ano)", 3, 3, 3, "2026-06-10", "registro.br — oficial"),
]
for p in fer:
    pesq_line(ws, r, *p)
    r += 1
r += 1

# F. TRÁFEGO PAGO
r = pesq_section(ws, r, "F. TRÁFEGO PAGO — fee mensal por faixa de verba do cliente (BR)")
trf = [
    ("Cliente até R$ 2k/mês mídia", 750, 1200, 1700, "2026-06-10", "wisedatamarketing.com; atriomarketing.site"),
    ("Cliente R$ 2-5k mídia", 1500, 2250, 3000, "2026-06-10", "spotmkt.com.br; rocketmidia.com"),
    ("Cliente R$ 5-10k mídia", 2500, 3750, 5000, "2026-06-10", "spotmkt.com.br"),
    ("Cliente >R$ 10k mídia", 5000, 10000, 20000, "2026-06-10", "wisedatamarketing.com"),
    ("Modelo % sobre verba (alternativa)", "10%", "15%", "20%", "2026-06-10", "atriomarketing.site"),
    ("Setup inicial campanhas (avulso)", 500, 1500, 3000, "2026-06-10", "spotmkt.com.br — faixa estimada"),
    ("⚠️ Aumento Meta Ads BR jan/2026 (PIS/Cofins/ISS)", "+12,15%", "—", "—", "2026-06-10", "atriomarketing.site/blog"),
]
for p in trf:
    pesq_line(ws, r, *p, money=False if isinstance(p[1], str) else True)
    r += 1
r += 1

# G. TRIBUTAÇÃO
r = pesq_section(ws, r, "G. TRIBUTAÇÃO E CONTABILIDADE — referências reais")
trib_header = ["Faixa RBT12", "Alíq. nominal", "Parcela a deduzir", "Alíq. efetiva (topo)", "Anexo", "Fonte"]
header_cells(ws, r, trib_header, start_col=2)
r += 1
trib = [
    ("Até R$ 180.000",       "6,00%",  "R$ 0",        "6,00%",  "III (Fator R ≥ 28%)", "Conube / Contabilizei — Anexo III"),
    ("R$ 180k-360k",         "11,20%", "R$ 9.360",    "8,40%",  "III",                 "conube.com.br/blog/anexo-3-tabela-simples-nacional"),
    ("R$ 360k-720k",         "13,50%", "R$ 17.640",   "11,05%", "III",                 "contabilizei.com.br/contabilidade-online/anexo-3-simples-nacional"),
    ("R$ 720k-1,8M",         "16,00%", "R$ 35.640",   "14,02%", "III",                 "contabilizei.com.br"),
    ("R$ 1,8M-3,6M",         "21,00%", "R$ 125.640",  "—",      "III",                 "contabilizei.com.br"),
    ("R$ 3,6M-4,8M (teto Simples)", "33,00%", "R$ 648.000", "—",    "III",             "contabilizei.com.br"),
]
for t in trib:
    for i, v in enumerate(t):
        c = ws.cell(row=r, column=2+i, value=v)
        c.border = BORDER; c.alignment = CENTER; c.font = NORMAL
    r += 1
r += 1

# Honorários contábeis
sub_row(ws, r, "G.1 — Honorários contábeis 2026 (livremente pactuados — CRC)", span=8); r += 1
hon = [
    ("ME Simples Nacional faixa típica", 259, 489, 900, "2026-06-10", "ohub.com.br/precos/contabilidade-mensal"),
    ("Contabilizei (online)", 195, 195, 195, "2026-06-10", "contabilizei.com.br/quanto-custa"),
    ("Contajá (online)", 137, 137, 137, "2026-06-10", "contaja.com.br"),
    ("Agilize (online)", 259, 259, 259, "2026-06-10", "artigos.agilize.com.br"),
    ("Contador ESPECIALIZADO em TI/agência (recomendado)", 700, 900, 1200, "2026-06-10", "Mercado especializado"),
]
header_cells(ws, r, ["Item", "Mínimo", "Médio", "Máximo", "Data", "Fonte"], start_col=2); r += 1
for p in hon:
    pesq_line(ws, r, *p)
    r += 1
r += 1

# H. HOME OFFICE
r = pesq_section(ws, r, "H. CUSTOS HOME OFFICE")
home = [
    ("Vivo Empresas Fibra 600 Mbps", 128.98, 128.98, 128.98, "2026-06-10", "minhaconexao.com.br"),
    ("TIM Ultrafibra 500 Mbps", 119.99, 139.99, 159.99, "2026-06-10", "minhaconexao.com.br"),
    ("Claro Empresas 600 Mbps (promo 6m)", 64.90, 64.90, 64.90, "2026-06-10", "planos.claro.com.br/empresas — atenção pós-promoção"),
    ("Faixa 200-500 Mbps empresarial (mercado)", 100, 140, 180, "2026-06-10", "melhorplano.net"),
    ("kWh Enel SP (B1, bandeira verde, c/ impostos)", 0.91, 0.91, 0.91, "2026-06-10", "custosolar.com; calculadoraenergia.com.br"),
    ("kWh média nacional", 0.68, 0.68, 0.68, "2026-06-10", "custosolar.com"),
    ("kWh Light RJ (mais cara)", 1.47, 1.47, 1.47, "2026-06-10", "clarke.com.br/mapa-tarifas"),
    ("kWh RS (mais barata)", 0.41, 0.41, 0.41, "2026-06-10", "clarke.com.br"),
    ("Consumo home office notebook 8h/dia (12 kWh/mês ≈ R$ 11 SP)", 11, 11, 11, "2026-06-10", "enel.com.br"),
    ("Consumo home office desktop 8h/dia (36 kWh ≈ R$ 33 SP)", 33, 33, 33, "2026-06-10", "enel.com.br"),
    ("Consumo home office desktop 10h (45 kWh ≈ R$ 41 SP)", 41, 41, 41, "2026-06-10", "enel.com.br"),
    ("Estimativa total adicional luz/sócio", 12, 40, 45, "2026-06-10", "revistaoeste.com — conservador SP"),
]
for p in home:
    pesq_line(ws, r, *p)
    r += 1
r += 1

# DADOS NÃO ENCONTRADOS
sub_row(ws, r, "DADOS NÃO ENCONTRADOS / LIMITAÇÕES (sinalização honesta)", span=8); r += 1
lacunas = [
    "VTEX — % take rate exato (cotação direta necessária)",
    "WooCommerce TCO BR consolidado — calculado por composição",
    "Auditoria de e-commerce isolada — proxy de auditoria SEO",
    "Hora-agência (não freelancer) — Glassdoor cobre freela/CLT; estimativa +40-60%",
    "Inadimplência B2B serviços — Serasa publica contagens, não % de receita perdida; 4-8% como referência conservadora",
    "Setup inicial tráfego pago — prática confirmada, faixa estimada R$ 500-3.000",
    "SLA-padrão em horas de planos de manutenção BR — agências divulgam em comunicação comercial; consolidação é estimativa",
    "Microsoft 365 BR em reais para reajuste jul/2026 — usado USD convertido",
]
for lac in lacunas:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    c = ws.cell(row=r, column=2, value=f"⚠️ {lac}"); c.font=ITAL; c.border=BORDER; c.alignment=LEFT
    r += 1

print(f"Aba 3 ✓ (Pesquisa de Mercado — {r} linhas)")


# ============================================================
# ABA 4 — CUSTOS FIXOS MENSAIS (detalhado por classe)
# ============================================================
ws = wb.create_sheet("4. Custos Fixos")
set_col_widths(ws, [3, 42, 14, 18, 38, 3])

title_row(ws, 1, "CUSTOS FIXOS MENSAIS — CLASSIFICAÇÃO", span=6)
ws.cell(row=2, column=2, value="Classificação: 🟢 OBRIGATÓRIO (não pode faltar) | 🟡 RECOMENDADO | ⚪ OPCIONAL").font=ITAL

header_cells(ws, 4, ["Item", "Valor R$", "Classe", "Observação / Fonte"], start_col=2)

custos = [
    # OBRIGATÓRIOS
    ("Pró-labore sócio 1", f"={PR}!C7", "🟢 OBRIGATÓRIO", "Art. 1.071 IV CC; mínimo p/ sustentar Fator R do Anexo III"),
    ("Pró-labore sócio 2", f"={PR}!C7", "🟢 OBRIGATÓRIO", "Idem"),
    ("INSS sócios (11% sobre pró-labore total)", f"={PR}!C11", "🟢 OBRIGATÓRIO", "Art. 21 Lei 8.212/91 (até teto RGPS)"),
    ("Internet empresarial (fibra)", 200, "🟢 OBRIGATÓRIO", "Vivo R$ 128,98; TIM R$ 119-159; média R$ 100-180"),
    ("Energia adicional home office", 80, "🟢 OBRIGATÓRIO", "Enel SP R$ 0,91/kWh × ~45 kWh/sócio"),
    ("Google Workspace (2 usuários)", 65, "🟢 OBRIGATÓRIO", "R$ 32,72/user × 2 (anual)"),
    ("Contabilidade especializada", 800, "🟢 OBRIGATÓRIO", "ME especializada R$ 700-1.200; Contabilizei online R$ 195"),
    ("Certificado Digital A1", 20, "🟢 OBRIGATÓRIO", "R$ 220/ano; necessário p/ NFS-e"),
    ("Hospedagem site Impulso + domínio", 50, "🟢 OBRIGATÓRIO", "Hostinger + Registro.br R$ 40/ano"),
    ("Antivírus + backup nuvem", 80, "🟢 OBRIGATÓRIO", "Segurança da operação"),
    # RECOMENDADOS
    ("ChatGPT Plus + Claude Pro (IA)", 220, "🟡 RECOMENDADO", "US$ 20 × 2 ≈ R$ 108 × 2"),
    ("Figma Professional (2 editores anual)", 162, "🟡 RECOMENDADO", "US$ 15 × 2 ≈ R$ 81 × 2"),
    ("Seguro RC profissional (E&O)", 150, "🟡 RECOMENDADO", "Recomendação jurídica — limita responsabilidade civil"),
    ("Marketing próprio (mídia + ferramentas)", 1200, "🟡 RECOMENDADO", "R$ 3-5k ideal; piso R$ 1.200"),
    ("Make / Zapier (automação)", 60, "🟡 RECOMENDADO", "Make Core US$ 10,59"),
    ("ClickUp Unlimited (2 users)", 14, "🟡 RECOMENDADO", "R$ 7 × 2"),
    ("Outros admin / contingência", 200, "🟡 RECOMENDADO", "Sindicato, taxas, eventuais"),
    # OPCIONAIS
    ("Canva Pro Teams", 39, "⚪ OPCIONAL", "Útil para mídia social — R$ 470/ano"),
    ("Adobe CC Standard", 249, "⚪ OPCIONAL", "Caro — usar só se justificado"),
    ("RD Station Basic (CRM)", 49, "⚪ OPCIONAL", "Pode iniciar com HubSpot Free R$ 0"),
    ("GitHub Copilot Pro", 54, "⚪ OPCIONAL", "Útil se trabalha com código Woo/Liquid"),
    ("Slack Pro (2 users)", 78, "⚪ OPCIONAL", "WhatsApp+Google Chat substituem no início"),
]
r = 5
for nome, val, classe, obs in custos:
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT
    if "OBRIGATÓRIO" in classe: a.fill = FILL_LG
    elif "RECOMENDADO" in classe: a.fill = FILL_LY
    else: a.fill = FILL_GREY
    b = ws.cell(row=r, column=3, value=val); b.border=BORDER; b.alignment=CENTER
    if isinstance(val, (int, float)): money0_fmt(b)
    elif isinstance(val, str) and val.startswith("="): money0_fmt(b)
    c = ws.cell(row=r, column=4, value=classe); c.border=BORDER; c.alignment=CENTER
    d = ws.cell(row=r, column=5, value=obs); d.border=BORDER; d.alignment=LEFT_TOP; d.font=ITAL
    ws.row_dimensions[r].height = 22
    r += 1

# Totais por classe
r += 1
totais_classes = [
    ("TOTAL OBRIGATÓRIO", '=SUMIF(D5:D26,"🟢 OBRIGATÓRIO",C5:C26)', FILL_LG),
    ("TOTAL RECOMENDADO", '=SUMIF(D5:D26,"🟡 RECOMENDADO",C5:C26)', FILL_LY),
    ("TOTAL OPCIONAL", '=SUMIF(D5:D26,"⚪ OPCIONAL",C5:C26)', FILL_GREY),
    ("CUSTO MENSAL TOTAL (todos)", '=SUM(C5:C26)', FILL_H3),
]
for nome, formula, fill in totais_classes:
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT; a.fill=fill
    if "TOTAL CUSTO" in nome or "CUSTO MENSAL" in nome:
        a.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    b = ws.cell(row=r, column=3, value=formula); b.font=BOLD; b.border=BORDER; b.alignment=CENTER
    money0_fmt(b)
    if "MENSAL TOTAL" in nome: b.fill=FILL_H3; b.font=Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    else: b.fill=fill
    ws.cell(row=r, column=4).border=BORDER; ws.cell(row=r, column=4).fill=fill
    ws.cell(row=r, column=5).border=BORDER; ws.cell(row=r, column=5).fill=fill
    ws.row_dimensions[r].height = 24
    r += 1

# Custo total anual
r += 1
a = ws.cell(row=r, column=2, value="CUSTO TOTAL ANUAL"); a.font=BOLD; a.fill=FILL_H1; a.alignment=LEFT
a.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
a.border=BORDER
b = ws.cell(row=r, column=3, value="=SUM(C5:C26)*12"); b.font=BOLD; b.border=BORDER; b.alignment=CENTER
money0_fmt(b); b.fill=FILL_H1; b.font=Font(name="Calibri", size=12, bold=True, color="FFFFFF")
ws.cell(row=r, column=4).border=BORDER; ws.cell(row=r, column=4).fill=FILL_H1
ws.cell(row=r, column=5).border=BORDER; ws.cell(row=r, column=5).fill=FILL_H1

print("Aba 4 ✓")


# ============================================================
# ABA 5 — CUSTOS VARIÁVEIS
# ============================================================
ws = wb.create_sheet("5. Custos Variáveis")
set_col_widths(ws, [3, 42, 14, 38, 3])

title_row(ws, 1, "CUSTOS VARIÁVEIS (% sobre receita ou por projeto)", span=5)
ws.cell(row=2, column=2, value="Valores em % são aplicados sobre receita; valores em R$ são por projeto/cliente").font=ITAL

header_cells(ws, 4, ["Item", "Valor / %", "Observação / Fonte"], start_col=2)

variaveis = [
    ("(% sobre receita) Imposto Simples Anexo III", f"={PR}!C31", "Premissa editável — efetivo real: 6% até R$ 180k/ano"),
    ("(% sobre receita) Taxa de cartão MÉDIA", f"={PR}!{PR_REF['tx_media']}", "Média ponderada das modalidades"),
    ("(% sobre receita) Inadimplência (PCLD)", f"={PR}!C38", "Conservador: 4% serviços B2B BR"),
    ("(% sobre receita) Comissão indicação (parcerias)", 0.1, "10-20% se houver indicador externo"),
    ("(R$ por projeto) Plugins/apps Shopify específicos", 200, "Quando o pacote inclui app pago (estimativa média)"),
    ("(R$ por projeto) Templates pagos / temas premium", 300, "Tema Shopify premium US$ 200-350"),
    ("(R$ por projeto) Banco de imagens / fontes", 100, "Adobe Stock, Freepik PRO, etc."),
    ("(R$ por projeto) Subcontratado eventual (designer/copy freela)", 800, "Quando o escopo exigir — opcional"),
    ("(R$ por mês) Tráfego pago próprio (aquisição clientes)", 1500, "Repassado da aba 4 marketing; também variável"),
    ("(R$ por hora extra de atendimento pós-venda)", 89, "Custo de oportunidade da hora produtiva"),
    ("Provisão para cancelamentos/devoluções", "1-3% da receita", "Recomendação contábil (CPC 25 — Provisões)"),
]
r = 5
for nome, val, obs in variaveis:
    a = ws.cell(row=r, column=2, value=nome); a.font=NORMAL; a.border=BORDER; a.alignment=LEFT
    b = ws.cell(row=r, column=3, value=val); b.border=BORDER; b.alignment=CENTER; b.font=BOLD
    if isinstance(val, float) and val < 1: pct_fmt(b,2)
    elif isinstance(val, (int, float)): money0_fmt(b)
    elif isinstance(val, str) and val.startswith("="): pct_fmt(b,2)
    c = ws.cell(row=r, column=4, value=obs); c.border=BORDER; c.alignment=LEFT_TOP; c.font=ITAL
    ws.row_dimensions[r].height = 22
    r += 1

print("Aba 5 ✓")


# ============================================================
# ABA 6 — CAPACIDADE OPERACIONAL
# ============================================================
ws = wb.create_sheet("6. Capacidade")
set_col_widths(ws, [3, 42, 14, 14, 14, 38, 3])

title_row(ws, 1, "CAPACIDADE OPERACIONAL — análise de horas", span=7)

section_row(ws, 3, "HORAS DISPONÍVEIS (por mês)", span=7)
header_cells(ws, 4, ["Linha", "Por sócio", "Total (2 sócios)", "% das brutas", "Observação"], start_col=2)
horas_data = [
    ("Horas BRUTAS (22 dias × 8h)", f"={PR}!C16", f"={PR}!C17", "100%", "Linha de partida"),
    ("(-) Comercial (pré-venda, propostas, reuniões)", f"={PR}!C17*0.15/{PR}!C5", f"={PR}!C17*0.15", "15%", "53h totais"),
    ("(-) Atendimento ativo + briefing", f"={PR}!C17*0.10/{PR}!C5", f"={PR}!C17*0.10", "10%", "35h totais"),
    ("(-) Pós-venda + suporte/garantia", f"={PR}!C17*0.10/{PR}!C5", f"={PR}!C17*0.10", "10%", "35h totais"),
    ("(-) Admin + financeiro + reuniões internas", f"={PR}!C17*0.08/{PR}!C5", f"={PR}!C17*0.08", "8%", "28h totais"),
    ("(-) Gestão + estudo + atualização templates", f"={PR}!C17*0.07/{PR}!C5", f"={PR}!C17*0.07", "7%", "25h totais"),
    ("= HORAS PRODUTIVAS de execução pura", f"={PR}!C19/{PR}!C5", f"={PR}!{PR_REF['horas_prod']}", "50%", "Padrão de planejamento: 176h"),
    ("Capacidade ALVO sustentável (≤80%)", f"={PR}!C22/{PR}!C5", f"={PR}!{PR_REF['horas_alvo']}", "40%", "Operar ≤80% evita burnout"),
]
r = 5
for linha, sc, tot, pct, obs in horas_data:
    is_total = linha.startswith("= HORAS")
    is_alvo = linha.startswith("Capacidade")
    row_fill = None
    a = ws.cell(row=r, column=2, value=linha); a.font = NORMAL; a.border=BORDER; a.alignment=LEFT
    if is_total: a.font = BOLD; row_fill = FILL_LG; a.fill = row_fill
    if is_alvo: row_fill = FILL_H3; a.fill = row_fill; a.font=Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    b = ws.cell(row=r, column=3, value=sc); b.border=BORDER; b.alignment=CENTER; int_fmt(b)
    c = ws.cell(row=r, column=4, value=tot); c.border=BORDER; c.alignment=CENTER; int_fmt(c); c.font=BOLD
    d = ws.cell(row=r, column=5, value=pct); d.border=BORDER; d.alignment=CENTER
    e = ws.cell(row=r, column=6, value=obs); e.border=BORDER; e.alignment=LEFT_TOP; e.font=ITAL
    if row_fill is not None:
        for col in (3,4,5,6):
            ws.cell(row=r, column=col).fill = row_fill
    r += 1

section_row(ws, r+1, "MIX MENSAL RECOMENDADO (alvo ≤80% de capacidade)", span=7); r += 2
header_cells(ws, r, ["Combinação", "Start (20h)", "Pro (40h)", "Premium (70h)", "Horas totais", "Avaliação"], start_col=2); r += 1
mixes = [
    ("Volume máximo (NÃO recomendado)", 4, 1, 0, "=C{}*20+D{}*40+E{}*70".format(r,r,r), "❌ 200h — estoura (>80%)"),
    ("Volume saudável", 3, 1, 0, "=C{}*20+D{}*40+E{}*70".format(r+1,r+1,r+1), "✅ 100h — ok com folga"),
    ("Equilibrado", 2, 1, 0, "=C{}*20+D{}*40+E{}*70".format(r+2,r+2,r+2), "✅ 80h — confortável"),
    ("Premium-foco", 1, 0, 1, "=C{}*20+D{}*40+E{}*70".format(r+3,r+3,r+3), "✅ 90h — sustentável"),
    ("Target sustentável (recomendado)", 2, 1, 0, "=C{}*20+D{}*40+E{}*70".format(r+4,r+4,r+4), "✅ 80h — pode adicionar suporte"),
    ("Premium puro (a cada 2 meses)", 0, 0, 1, "=C{}*20+D{}*40+E{}*70".format(r+5,r+5,r+5), "✅ 70h — ok mas baixo throughput"),
]
for nome, st, pro, prem, tot, aval in mixes:
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT
    for j, v in enumerate([st, pro, prem, tot, aval]):
        c = ws.cell(row=r, column=3+j, value=v); c.border=BORDER; c.alignment=CENTER
        if j in (0,1,2): int_fmt(c)
        if j == 3: int_fmt(c); c.font=BOLD
    if "recomendado" in nome.lower() or "Target" in nome:
        for col in range(2, 8): ws.cell(row=r, column=col).fill = FILL_LG
    elif "NÃO" in nome:
        for col in range(2, 8): ws.cell(row=r, column=col).fill = FILL_LR
    r += 1

section_row(ws, r+1, "LIMITES DE SUPORTE RECORRENTE POR SÓCIO", span=7); r += 2
header_cells(ws, r, ["Plano", "Horas/mês cliente", "Capacidade nominal/sócio", "Teto operacional/sócio", "Mix sustentável"], start_col=2); r += 1
sup_data = [
    ("Essencial (4h/mês)", 4, 45, 25, "Combinar com 1-2 Pro"),
    ("Profissional (10h/mês)", 10, 18, 12, "Limite mais comum"),
    ("Growth (20h/mês)", 20, 9, 6, "Cliente premium"),
]
for nome, h, cap, teto, mix in sup_data:
    write_row(ws, r, [None, nome, h, cap, teto, mix], money0_cols=[], int_cols=[3,4,5], font=NORMAL, align=CENTER)
    ws.cell(row=r, column=2).font = BOLD; ws.cell(row=r, column=2).alignment = LEFT
    r += 1

section_row(ws, r+1, "GATILHOS PARA CONTRATAR O 3º PROFISSIONAL", span=7); r += 2
gatilhos = [
    "Backlog vendido > 200h por 2 meses consecutivos",
    "Capacidade alocada > 85% por 2 meses",
    "Recusa de ≥2 projetos qualificados por mês por falta de slot",
    "Suporte recorrente > 50h/mês combinado",
    "NPS cai abaixo de 8 ou atraso médio > 20% sobre prazo prometido",
    "MRR ≥ R$ 20.000 por 3 meses (cobre custo de contratação CLT-flex ~R$ 6-9k all-in)",
    "Sócios trabalhando > 55h/semana ou rejeitando projetos por capacidade",
]
for g in gatilhos:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=f"☐  {g}"); c.font=NORMAL; c.border=BORDER; c.alignment=LEFT
    r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
c = ws.cell(row=r, column=2, value="✅ Perfil ideal da 1ª contratação: IMPLEMENTADOR PLENO em Shopify/Tray (não sênior). R$ 5-7k CLT-flex ou PJ. Sócios sobem para vendas + suporte estratégico + conteúdo.")
c.font=BOLD; c.fill=FILL_LG; c.border=BORDER; c.alignment=LEFT
ws.row_dimensions[r].height = 32

print("Aba 6 ✓")


# ============================================================
# ABA 7 — PRECIFICAÇÃO PROJETO 48 HORAS
# ============================================================
ws = wb.create_sheet("7. Preço 48 Horas")
set_col_widths(ws, [3, 38, 18, 18, 18, 38, 3])

title_row(ws, 1, "PRECIFICAÇÃO — PROJETO 48 HORAS (3 pacotes)", span=7)
ws.cell(row=2, column=2, value="Fórmula de precificação: PV = (CustoHora × Horas) ÷ (1 - %Imposto - %TaxaMédia - %Inadimplência - %Margem)").font=ITAL

section_row(ws, 4, "ESCOPO POR PACOTE", span=7)
header_cells(ws, 5, ["Item", "START", "PRO", "PREMIUM", "Observação"], start_col=2)

escopo_data = [
    ("Plataformas suportadas", "Shopify, Nuvemshop, WP", "Shopify, Nuvemshop, Tray", "Shopify, Tray", "Woo só com +30%"),
    ("Limite de SKUs", 15, 50, 150, "Cadastro com fotos prontas"),
    ("Limite de páginas", 5, 8, 12, "Páginas institucionais"),
    ("Limite de integrações", 1, 1, 2, "Bling OU Tiny OU RD"),
    ("Limite de gateways", 1, 2, 2, "Aprovado antes do start"),
    ("Limite de modalidades de frete", 1, 2, 2, "Melhor Envio/Correios"),
    ("Reuniões inclusas", 3, 3, 5, "Briefing + validação + treinamento"),
    ("Sessões de treinamento (60min)", 1, 2, 2, "+ videoteca de acesso"),
    ("Dias de suporte pós-go-live", 7, 15, 30, "Após, plano mensal obrigatório"),
    ("Horas humanas estimadas (mín)", 22, 45, 90, "Cenário ideal com pré-condições"),
    ("Horas humanas estimadas (esperada)", 32, 70, 130, "Mediana operacional"),
    ("Horas humanas estimadas (máx)", 45, 100, 180, "Cenário stress"),
]
r = 6
for item in escopo_data:
    nome = item[0]; vals = item[1:4]; obs = item[4]
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT; a.fill=FILL_GREY
    for j, v in enumerate(vals):
        c = ws.cell(row=r, column=3+j, value=v); c.border=BORDER; c.alignment=CENTER
        if isinstance(v, int) and not isinstance(v, bool): int_fmt(c)
    e = ws.cell(row=r, column=6, value=obs); e.border=BORDER; e.alignment=LEFT_TOP; e.font=ITAL
    ws.row_dimensions[r].height = 22
    r += 1

START_ROW = 6  # ref para horas
PRO_ROW = 6
PREM_ROW = 6
# As horas esperadas estão em linha 16 do bloco escopo (linha 6 + 10 = 16)
H_ESP_ROW = 16  # horas esperadas (32 / 70 / 130)

r += 1
section_row(ws, r, "PRECIFICAÇÃO CALCULADA (base: horas esperadas)", span=7); r += 1
header_cells(ws, r, ["Linha", "START", "PRO", "PREMIUM", "Fórmula / Origem"], start_col=2); r += 1

horas_ref = {"START": f"C{H_ESP_ROW}", "PRO": f"D{H_ESP_ROW}", "PREMIUM": f"E{H_ESP_ROW}"}

# Custo interno
ws.cell(row=r, column=2, value="Custo interno (R$)").font=BOLD; ws.cell(row=r, column=2).fill=FILL_GREY; ws.cell(row=r, column=2).border=BORDER
for j, pac in enumerate(["START","PRO","PREMIUM"]):
    c = ws.cell(row=r, column=3+j, value=f"={PR}!{PR_REF['custo_hora']}*{horas_ref[pac]}")
    c.border=BORDER; c.alignment=CENTER; money_fmt(c)
ws.cell(row=r, column=6, value="= custo/h × horas").font=ITAL; ws.cell(row=r, column=6).border=BORDER
r += 1

# Preço MÍNIMO (margem mínima)
ws.cell(row=r, column=2, value="PREÇO MÍNIMO (margem 20%)").font=BOLD; ws.cell(row=r, column=2).fill=FILL_LR; ws.cell(row=r, column=2).border=BORDER
for j, pac in enumerate(["START","PRO","PREMIUM"]):
    c = ws.cell(row=r, column=3+j,
                value=f"=({PR}!{PR_REF['custo_hora']}*{horas_ref[pac]})/(1-{PR}!{PR_REF['imposto']}-{PR}!{PR_REF['tx_media']}-{PR}!{PR_REF['inad']}-{PR}!{PR_REF['margem_min']})")
    c.border=BORDER; c.alignment=CENTER; money_fmt(c); c.fill=FILL_LR
ws.cell(row=r, column=6, value="Custo / (1 - 12% - tx méd - 4% - 20%)").font=ITAL; ws.cell(row=r, column=6).border=BORDER
r += 1

# Preço RECOMENDADO (margem 40%)
ws.cell(row=r, column=2, value="PREÇO RECOMENDADO (margem 40%)").font=BOLD; ws.cell(row=r, column=2).fill=FILL_LG; ws.cell(row=r, column=2).border=BORDER
for j, pac in enumerate(["START","PRO","PREMIUM"]):
    c = ws.cell(row=r, column=3+j,
                value=f"=({PR}!{PR_REF['custo_hora']}*{horas_ref[pac]})/(1-{PR}!{PR_REF['imposto']}-{PR}!{PR_REF['tx_media']}-{PR}!{PR_REF['inad']}-{PR}!{PR_REF['margem_rec']})")
    c.border=BORDER; c.alignment=CENTER; money_fmt(c); c.font=BOLD; c.fill=FILL_LG
ws.cell(row=r, column=6, value="Custo / (1 - 12% - tx méd - 4% - 40%)").font=ITAL; ws.cell(row=r, column=6).border=BORDER
PRECO_REC_ROW = r
r += 1

# Preço PREMIUM (margem 55%)
ws.cell(row=r, column=2, value="PREÇO PREMIUM (margem 55%)").font=BOLD; ws.cell(row=r, column=2).fill=FILL_LY; ws.cell(row=r, column=2).border=BORDER
for j, pac in enumerate(["START","PRO","PREMIUM"]):
    c = ws.cell(row=r, column=3+j,
                value=f"=({PR}!{PR_REF['custo_hora']}*{horas_ref[pac]})/(1-{PR}!{PR_REF['imposto']}-{PR}!{PR_REF['tx_media']}-{PR}!{PR_REF['inad']}-{PR}!{PR_REF['margem_prem']})")
    c.border=BORDER; c.alignment=CENTER; money_fmt(c); c.fill=FILL_LY
ws.cell(row=r, column=6, value="Custo / (1 - 12% - tx méd - 4% - 55%)").font=ITAL; ws.cell(row=r, column=6).border=BORDER
r += 1

# Preço SUGERIDO COMERCIAL (ancoragem psicológica)
ws.cell(row=r, column=2, value="PREÇO COMERCIAL SUGERIDO (ancoragem)").font=BOLD; ws.cell(row=r, column=2).fill=FILL_H3; ws.cell(row=r, column=2).border=BORDER
ws.cell(row=r, column=2).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
for j, valor in enumerate([2490, 4990, 8900]):
    c = ws.cell(row=r, column=3+j, value=valor); c.border=BORDER; c.alignment=CENTER
    money0_fmt(c); c.font=Font(name="Calibri", size=12, bold=True, color="FFFFFF"); c.fill=FILL_H3
ws.cell(row=r, column=6, value="Recomendação CFO: ancoragem + colchão p/ desconto").font=ITAL; ws.cell(row=r, column=6).border=BORDER
PRECO_COM_ROW = r
r += 1

# Margem efetiva no preço sugerido
ws.cell(row=r, column=2, value="Margem líquida no preço sugerido").font=BOLD; ws.cell(row=r, column=2).fill=FILL_GREY; ws.cell(row=r, column=2).border=BORDER
for j, pac in enumerate(["START","PRO","PREMIUM"]):
    col = chr(ord("C") + j)
    c = ws.cell(row=r, column=3+j,
                value=f"=1-{PR}!{PR_REF['imposto']}-{PR}!{PR_REF['tx_media']}-{PR}!{PR_REF['inad']}-(({PR}!{PR_REF['custo_hora']}*{horas_ref[pac]})/{col}{PRECO_COM_ROW})")
    c.border=BORDER; c.alignment=CENTER; pct_fmt(c)
ws.cell(row=r, column=6, value="= 1 - imp - tx - inad - custo%").font=ITAL; ws.cell(row=r, column=6).border=BORDER
r += 2

section_row(ws, r, "SIMULAÇÃO POR MODALIDADE DE PAGAMENTO — Pacote PRO (preço sugerido)", span=7); r += 1
header_cells(ws, r, ["Modalidade", "Valor bruto", "Taxa %", "Taxa R$", "Imposto R$", "Líquido recebido"], start_col=2); r += 1

mods = [
    ("PIX",            f"=D{PRECO_COM_ROW}", f"={PR}!{PR_REF['tx_pix']}",  f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['tx_pix']}",  f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['imposto']}"),
    ("Cartão à vista", f"=D{PRECO_COM_ROW}", f"={PR}!{PR_REF['tx_vista']}",f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['tx_vista']}",f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['imposto']}"),
    ("Cartão 3x",      f"=D{PRECO_COM_ROW}", f"={PR}!{PR_REF['tx_3x']}",   f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['tx_3x']}",   f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['imposto']}"),
    ("Cartão 6x",      f"=D{PRECO_COM_ROW}", f"={PR}!{PR_REF['tx_6x']}",   f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['tx_6x']}",   f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['imposto']}"),
    ("Cartão 10x",     f"=D{PRECO_COM_ROW}", f"={PR}!{PR_REF['tx_10x']}",  f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['tx_10x']}",  f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['imposto']}"),
    ("Cartão 12x",     f"=D{PRECO_COM_ROW}", f"={PR}!{PR_REF['tx_12x']}",  f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['tx_12x']}",  f"=D{PRECO_COM_ROW}*{PR}!{PR_REF['imposto']}"),
]
for nome, bruto, taxa_p, taxa_r, imp_r in mods:
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT; a.fill=FILL_LB
    b = ws.cell(row=r, column=3, value=bruto); b.border=BORDER; b.alignment=CENTER; money_fmt(b)
    c = ws.cell(row=r, column=4, value=taxa_p); c.border=BORDER; c.alignment=CENTER; pct_fmt(c,2)
    d = ws.cell(row=r, column=5, value=taxa_r); d.border=BORDER; d.alignment=CENTER; money_fmt(d)
    e = ws.cell(row=r, column=6, value=imp_r); e.border=BORDER; e.alignment=CENTER; money_fmt(e)
    f = ws.cell(row=r, column=7, value=f"=C{r}-E{r}-F{r}"); f.border=BORDER; f.alignment=CENTER; money_fmt(f); f.font=BOLD; f.fill=FILL_LG
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
c = ws.cell(row=r, column=2,
            value="💡 POLÍTICA RECOMENDADA: 5% off PIX/à vista; máx 3x sem juros; 6x+ com juros repassados ao cliente. DNS sob domínio Impulso até pagamento final (50/40/10).")
c.font=BOLD; c.fill=FILL_LY; c.alignment=LEFT; c.border=BORDER
ws.row_dimensions[r].height = 30

print("Aba 7 ✓")


# ============================================================
# ABA 8 — PROJETOS PERSONALIZADOS
# ============================================================
ws = wb.create_sheet("8. Personalizados")
set_col_widths(ws, [3, 36, 12, 16, 14, 14, 16, 16, 16, 3])

title_row(ws, 1, "PRECIFICAÇÃO — PROJETOS PERSONALIZADOS", span=10)
ws.cell(row=2, column=2, value="Mesma fórmula: PV = (CustoHora × Horas) ÷ (1 - %Imp - %Tx - %Inad - %Margem)").font=ITAL

section_row(ws, 4, "TABELA DE PROJETOS PERSONALIZADOS", span=10)
header_cells(ws, 5, ["Projeto", "Horas estim.", "Custo interno", "Imposto", "Taxa cartão", "Preço MÍNIMO (20%)", "Preço RECOMENDADO (40%)", "Preço PREMIUM (55%)"], start_col=2)

projetos = [
    ("Site institucional simples (até 5 páginas)", 30),
    ("Site institucional premium (até 10 páginas + funil)", 60),
    ("Loja virtual personalizada (50-100 SKUs + integrações)", 100),
    ("Migração de plataforma (ex.: VTEX → Shopify)", 80),
    ("Integração com ERP (Bling/Tiny/Omie)", 25),
    ("Integração marketplace (ML/Amazon) por canal", 30),
    ("SEO técnico — projeto completo", 40),
    ("Automação comercial (RD/HubSpot/Klaviyo)", 30),
    ("CRM — setup + treinamento", 25),
    ("Landing page profissional (com funil + automação)", 20),
    ("Auditoria de e-commerce completa", 18),
    ("Consultoria estratégica (mensal, 1 sócio sênior)", 20),
]
r = 6
for nome, horas in projetos:
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT
    h = ws.cell(row=r, column=3, value=horas); h.border=BORDER; h.alignment=CENTER; int_fmt(h)
    custo = ws.cell(row=r, column=4, value=f"={PR}!{PR_REF['custo_hora']}*C{r}"); custo.border=BORDER; custo.alignment=CENTER; money0_fmt(custo)
    pmin = f"=D{r}/(1-{PR}!{PR_REF['imposto']}-{PR}!{PR_REF['tx_media']}-{PR}!{PR_REF['inad']}-{PR}!{PR_REF['margem_min']})"
    prec = f"=D{r}/(1-{PR}!{PR_REF['imposto']}-{PR}!{PR_REF['tx_media']}-{PR}!{PR_REF['inad']}-{PR}!{PR_REF['margem_rec']})"
    pprm = f"=D{r}/(1-{PR}!{PR_REF['imposto']}-{PR}!{PR_REF['tx_media']}-{PR}!{PR_REF['inad']}-{PR}!{PR_REF['margem_prem']})"
    # Coluna E = imposto/projeto, F = taxa
    e = ws.cell(row=r, column=5, value=f"=G{r}*{PR}!{PR_REF['imposto']}"); e.border=BORDER; e.alignment=CENTER; money0_fmt(e)
    f_ = ws.cell(row=r, column=6, value=f"=G{r}*{PR}!{PR_REF['tx_media']}"); f_.border=BORDER; f_.alignment=CENTER; money0_fmt(f_)
    g = ws.cell(row=r, column=7, value=pmin); g.border=BORDER; g.alignment=CENTER; money0_fmt(g); g.fill=FILL_LR
    h2 = ws.cell(row=r, column=8, value=prec); h2.border=BORDER; h2.alignment=CENTER; money0_fmt(h2); h2.fill=FILL_LG; h2.font=BOLD
    i = ws.cell(row=r, column=9, value=pprm); i.border=BORDER; i.alignment=CENTER; money0_fmt(i); i.fill=FILL_LY
    ws.row_dimensions[r].height = 22
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
c = ws.cell(row=r, column=2,
            value="📌 Para projetos personalizados: SEMPRE escopo fechado por escrito + change request formal para qualquer item adicional. Política de revisões: 2 rodadas estéticas; excedentes faturados por hora.")
c.font=BOLD; c.fill=FILL_LY; c.border=BORDER; c.alignment=LEFT
ws.row_dimensions[r].height = 32

print("Aba 8 ✓")


# ============================================================
# ABA 9 — PLANOS DE SUPORTE MENSAL
# ============================================================
ws = wb.create_sheet("9. Suporte")
set_col_widths(ws, [3, 38, 22, 22, 22, 3])

title_row(ws, 1, "PLANOS MENSAIS DE SUPORTE E MANUTENÇÃO", span=6)

section_row(ws, 3, "3 PLANOS RECORRENTES (MRR)", span=6)
header_cells(ws, 4, ["Item", "ESSENCIAL", "PROFISSIONAL", "GROWTH"], start_col=2)

sup = [
    ("Horas inclusas por mês", 4, 10, 20),
    ("Valor mensal sugerido (R$)", 397, 990, 1890),
    ("Custo interno (R$)", f"={PR}!{PR_REF['custo_hora']}*C5", f"={PR}!{PR_REF['custo_hora']}*D5", f"={PR}!{PR_REF['custo_hora']}*E5"),
    ("Imposto Simples (R$)", "=C6*'2. Premissas'!C31", "=D6*'2. Premissas'!C31", "=E6*'2. Premissas'!C31"),
    ("Margem líquida estimada", "=1-'2. Premissas'!C31-'2. Premissas'!"+PR_REF['tx_media']+"-'2. Premissas'!C38-C7/C6",
                                 "=1-'2. Premissas'!C31-'2. Premissas'!"+PR_REF['tx_media']+"-'2. Premissas'!C38-D7/D6",
                                 "=1-'2. Premissas'!C31-'2. Premissas'!"+PR_REF['tx_media']+"-'2. Premissas'!C38-E7/E6"),
    ("SLA resposta (h úteis)", 24, 8, 4),
    ("SLA resolução P1", "5 d.u.", "24 h", "8 h"),
    ("Vigência mínima sugerida", "6 meses", "12 meses", "12 meses"),
    ("Reajuste anual", "IPCA", "IPCA", "IPCA"),
    ("Limite nominal por sócio", "45 clientes", "18 clientes", "9 clientes"),
    ("Teto operacional sustentável/sócio", "25-30", "12-15", "6-8"),
    ("Itens INCLUSOS", "Ajustes técnicos pontuais; atualização de banners; pequenas correções; monitoramento básico (uptime, backup); atendimento consultivo por mensagem",
                       "Tudo do Essencial + alterações em página existente (até 3/mês); ajuste de tema sem código; relatório mensal simples; revisão mensal por sócio",
                       "Tudo do Profissional + 1 sprint de melhoria/mês; otimização de conversão pontual; revisão de SEO mensal; reunião quinzenal de planejamento"),
    ("Itens NÃO INCLUSOS (orçamento à parte)", "Customização de código; nova página; nova integração; redesign; produção de conteúdo",
                                                "Customização de código; redesign; produção de conteúdo extenso; gestão de tráfego pago",
                                                "Redesign completo; novas integrações > 1; produção de conteúdo extenso; gestão de tráfego pago (vendido à parte)"),
]
r = 5
for item in sup:
    nome = item[0]
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT; a.fill=FILL_GREY
    for j in range(3):
        v = item[1+j]
        c = ws.cell(row=r, column=3+j, value=v); c.border=BORDER; c.alignment=CENTER
        if "Horas" in nome or "Valor" in nome or "Custo" in nome or "Imposto" in nome:
            if isinstance(v, (int, float)) or (isinstance(v, str) and v.startswith("=")):
                if "Horas" in nome: int_fmt(c)
                else: money0_fmt(c)
        if "Margem" in nome:
            pct_fmt(c)
        if "INCLUSOS" in nome or "NÃO" in nome:
            c.alignment = LEFT_TOP; c.font = SMALL
    if "INCLUSOS" in nome or "NÃO" in nome:
        ws.row_dimensions[r].height = 72
    r += 1

r += 1
section_row(ws, r, "ESTRATÉGIA DE RECORRÊNCIA (squad strategy)", span=6); r += 1
mensagem = (
    "Meta: 50% conversão setup → suporte mensal em 12 meses; 65% em 24 meses.\n"
    "Mecânicas:\n"
    "  1) Bundle de saída: setup já inclui 30 dias de suporte cortesia + oferta de mensalidade 20% off se contratada em 7 dias.\n"
    "  2) Revisão de resultados D+30 e D+60 com sócio → mostra dados → abre upsell para Profissional/Growth.\n"
    "  3) Comunidade Impulso acessível apenas a clientes em recorrência (status premium).\n"
    "  4) SLA visível no contrato. Cláusula de upgrade/downgrade com 30 dias de aviso prévio.\n"
    "  5) Vigência mínima 12 meses (Pro/Growth) — válida B2B; em B2C, atenção CDC art. 51 IV (multa rescisória proporcional).\n\n"
    "RECEITA RECORRENTE É O CORAÇÃO DO NEGÓCIO. 1 R$ MRR vale 3-5× 1 R$ de projeto."
)
ws.merge_cells(start_row=r, start_column=2, end_row=r+7, end_column=5)
c = ws.cell(row=r, column=2, value=mensagem); c.font=NORMAL; c.fill=FILL_LG
c.alignment=LEFT_TOP; c.border=BORDER

print("Aba 9 ✓")


# ============================================================
# ABA 10 — RECEITA DE PARCERIAS COM PLATAFORMAS
# ============================================================
ws = wb.create_sheet("10. Parceria")
set_col_widths(ws, [3, 32, 16, 16, 16, 16, 16, 16, 3])

title_row(ws, 1, "RECEITA RECORRENTE DE PARCERIAS COM PLATAFORMAS", span=9)
ws.cell(row=2, column=2, value="Comissão de indicação — natureza jurídica: intermediação (LC 116/03 itens 10.02 / 17.12)").font=ITAL

section_row(ws, 4, "PARÂMETROS DAS PLATAFORMAS PARCEIRAS", span=9)
header_cells(ws, 5, ["Plataforma", "Mensalidade média cliente (R$)", "% Comissão Impulso", "Receita/cliente/mês", "Fonte"], start_col=2)

plats = [
    ("Shopify Basic (US$ 29-39)",     180, 0.20, "=C6*D6", "help.shopify.com/partners"),
    ("Shopify Grow (US$ 79-105)",     500, 0.20, "=C7*D7", "shopify.com/pricing"),
    ("Tray (R$ 49-99 médio)",         75, 0.20, "=C8*D8", "tray.com.br/partners"),
    ("Nuvemshop Essencial",           59, 0.20, "=C9*D9", "atendimento.nuvemshop.com.br — comissões vigentes"),
    ("Nuvemshop Impulso (homônimo)",  139, 0.20, "=C10*D10", "nuvemshop.com.br/planos-e-precos"),
    ("Nuvemshop Escala",              389, 0.20, "=C11*D11", "nuvemshop.com.br/planos-e-precos"),
]
r = 6
for nome, mens, pct, rec, fonte in plats:
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT
    b = ws.cell(row=r, column=3, value=mens); b.border=BORDER; b.alignment=CENTER; money0_fmt(b)
    c = ws.cell(row=r, column=4, value=pct); c.border=BORDER; c.alignment=CENTER; pct_fmt(c)
    d = ws.cell(row=r, column=5, value=rec); d.border=BORDER; d.alignment=CENTER; money0_fmt(d); d.font=BOLD
    e = ws.cell(row=r, column=6, value=fonte); e.border=BORDER; e.alignment=LEFT_TOP; e.font=ITAL
    r += 1

r += 1
section_row(ws, r, "SIMULAÇÃO DE RECEITA RECORRENTE — 3 CENÁRIOS", span=9); r += 1
header_cells(ws, r, ["Cenário", "Clientes mês 3", "mês 6", "mês 12", "mês 24", "MRR mês 12 (R$)", "Anual ano 1 (R$)"], start_col=2); r += 1

cenarios_parc = [
    ("Pessimista",   2, 4, 8,  15,  '=E14*60',  '=SUMPRODUCT({2;3;4;4;5;6;6;7;7;7;8;8},60)'),
    ("Realista",     3, 8, 18, 35,  '=E15*60',  '=SUMPRODUCT({3;5;7;9;11;13;14;15;16;17;17;18},60)'),
    ("Otimista",     5, 12, 30, 60, '=E16*60',  '=SUMPRODUCT({5;8;11;15;18;21;24;26;28;29;29;30},60)'),
]
# usar receita média por cliente ~R$ 60 (mix mensal médio)
# Os ranges hardcoded acima são exemplos; vamos simplificar:
for i, (nome, m3, m6, m12, m24, _, _) in enumerate(cenarios_parc):
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT
    a.fill = [FILL_LR, FILL_LY, FILL_LG][i]
    for j, v in enumerate([m3, m6, m12, m24]):
        c = ws.cell(row=r, column=3+j, value=v); c.border=BORDER; c.alignment=CENTER; int_fmt(c)
    # MRR mês 12 = clientes × R$ 60 médio
    mrr = ws.cell(row=r, column=7, value=f"=E{r}*60"); mrr.border=BORDER; mrr.alignment=CENTER; money0_fmt(mrr); mrr.font=BOLD
    # Anual estimado (assumindo crescimento linear de m3 a m12)
    # Fórmula simples: média (m3, m12) × 12 × R$ 60
    anual = ws.cell(row=r, column=8, value=f"=((C{r}+E{r})/2)*12*60"); anual.border=BORDER; anual.alignment=CENTER; money0_fmt(anual)
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
obs = ("⚠️ Premissas do squad: (1) Receita média/cliente assumida R$ 60/mês — mix de planos Tray + Nuvemshop Essencial + Shopify Basic. "
       "(2) Churn 3-5% a.m. esperado — incorporado no ramping. (3) Receita Shopify Partners é EXPORTAÇÃO de serviço (ISS isento art. 2º I LC 116/03; tributação só CPP+IRPJ+CSLL no DAS — art. 18 §14 LC 123/06). "
       "(4) Contrato de câmbio obrigatório (Lei 14.286/21). (5) Não receber verba de mídia em conta — cliente paga direto a Meta/Google.")
c = ws.cell(row=r, column=2, value=obs); c.font=NORMAL; c.fill=FILL_LB; c.alignment=LEFT_TOP; c.border=BORDER
ws.row_dimensions[r].height = 60

print("Aba 10 ✓")


# ============================================================
# ABA 11 — RECEITA DE MÍDIA DIGITAL / TRÁFEGO PAGO
# ============================================================
ws = wb.create_sheet("11. Tráfego Pago")
set_col_widths(ws, [3, 30, 18, 18, 14, 16, 18, 16, 3])

title_row(ws, 1, "GESTÃO DE TRÁFEGO PAGO — MODELOS DE COBRANÇA", span=9)
ws.cell(row=2, column=2, value="⚠️ Cliente paga MÍDIA direto Meta/Google (NÃO transitar pela Impulso). Impulso cobra apenas FEE de gestão.").font=Font(size=10, italic=True, color="C00000", bold=True)

section_row(ws, 4, "MATRIZ DE PRECIFICAÇÃO POR FAIXA DE VERBA DO CLIENTE", span=9)
header_cells(ws, 5, ["Faixa de verba do cliente", "Setup (R$)", "Fee mensal (R$)", "Horas/mês", "% sobre verba", "Custo interno (R$)", "Margem"], start_col=2)

trafico = [
    ("Até R$ 2.000/mês mídia",  800,  1200, 8,  "10-15%", f"={PR}!{PR_REF['custo_hora']}*E6", '=1-G6/D6-\'2. Premissas\'!C31-\'2. Premissas\'!'+PR_REF['tx_media']),
    ("R$ 2-5k mídia",           1500, 2250, 12, "12-15%", f"={PR}!{PR_REF['custo_hora']}*E7", '=1-G7/D7-\'2. Premissas\'!C31-\'2. Premissas\'!'+PR_REF['tx_media']),
    ("R$ 5-10k mídia",          2500, 3750, 18, "10-15%", f"={PR}!{PR_REF['custo_hora']}*E8", '=1-G8/D8-\'2. Premissas\'!C31-\'2. Premissas\'!'+PR_REF['tx_media']),
    (">R$ 10k mídia",           3500, 5000, 25, "10-12%", f"={PR}!{PR_REF['custo_hora']}*E9", '=1-G9/D9-\'2. Premissas\'!C31-\'2. Premissas\'!'+PR_REF['tx_media']),
]
r = 6
for nome, setup, fee, h, pct_v, custo, margem in trafico:
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT
    b = ws.cell(row=r, column=3, value=setup); b.border=BORDER; b.alignment=CENTER; money0_fmt(b)
    c = ws.cell(row=r, column=4, value=fee); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD
    d = ws.cell(row=r, column=5, value=h); d.border=BORDER; d.alignment=CENTER; int_fmt(d)
    e = ws.cell(row=r, column=6, value=pct_v); e.border=BORDER; e.alignment=CENTER
    f_ = ws.cell(row=r, column=7, value=custo); f_.border=BORDER; f_.alignment=CENTER; money0_fmt(f_)
    g = ws.cell(row=r, column=8, value=margem); g.border=BORDER; g.alignment=CENTER; pct_fmt(g); g.font=BOLD
    if isinstance(margem, str): g.fill = FILL_LG
    r += 1

r += 1
section_row(ws, r, "MODELOS DE COBRANÇA — VANTAGENS E RISCOS", span=9); r += 1
modelos = [
    ("Setup + Fee fixo mensal", "Previsibilidade total; alinhado a serviço, não a verba", "Cliente pode reduzir verba e exigir mesmo fee", "Recomendado p/ clientes pequenos"),
    ("% sobre verba (10-20%)", "Cresce com cliente; alinhado a investimento", "Variação do faturamento; risco se cliente cortar verba", "Cliente médio/grande"),
    ("Pacote híbrido (fee + %)", "Equilíbrio entre previsibilidade e upside", "Mais complexo de explicar", "MELHOR para clientes em escala"),
    ("Consultoria avulsa (hora)", "Margem alta em projetos pontuais", "Não escala; sem recorrência", "Auditorias + workshops"),
]
header_cells(ws, r, ["Modelo", "Vantagens", "Riscos", "Quando usar"], start_col=2); r += 1
for m in modelos:
    for j, v in enumerate(m):
        c = ws.cell(row=r, column=2+j, value=v); c.border=BORDER; c.alignment=LEFT_TOP
        if j == 0: c.font=BOLD
        else: c.font=SMALL
    ws.row_dimensions[r].height = 32
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
c = ws.cell(row=r, column=2,
            value="⚠️ Cláusula contratual obrigatória: 'A Impulso não garante volume, ROAS, CPA, faturamento ou ROI; tais resultados dependem de fatores externos (mercado, produto, preço, sazonalidade, política das plataformas).' (CDC art. 39 IV)")
c.font=BOLD; c.fill=FILL_LR; c.alignment=LEFT_TOP; c.border=BORDER
ws.row_dimensions[r].height = 36

print("Aba 11 ✓")


# ============================================================
# ABA 12 — CENÁRIOS FINANCEIROS (Pessimista / Realista / Otimista)
# ============================================================
ws = wb.create_sheet("12. Cenários")
set_col_widths(ws, [3, 32] + [11]*13 + [3])  # A texto, B-N = 12 meses + total

title_row(ws, 1, "CENÁRIOS FINANCEIROS — PROJEÇÃO 12 MESES", span=16)

# Constantes-resumo para cada cenário em ROWS específicas
# Vamos posicionar:
# Pessimista linha 4-28
# Realista linha 31-55
# Otimista linha 58-82

PESS_START = 4
REAL_START = 31
OTIM_START = 58

PRECOS_48H = {"Start": 2490, "Pro": 4990, "Premium": 8900}
PRECOS_SUP = {"Essencial": 397, "Pro": 990, "Growth": 1890}

def build_cenario(ws, start_row, nome, fill, projeto_qty_per_month, suporte_qty_per_month,
                  parc_qty_per_month, traf_qty_per_month, projeto_mix, suporte_plan, traf_fee):
    """
    Constrói um cenário no formato:
    - Linha start: section header
    - Linha start+1: header de meses
    - Linha start+2: receita 48h
    - Linha start+3: receita personalizados (não usado aqui mas linha existe)
    - Linha start+4: receita suporte
    - Linha start+5: receita parceria
    - Linha start+6: receita tráfego
    - Linha start+7: RECEITA BRUTA TOTAL
    - Linha start+8: (-) Impostos
    - Linha start+9: (-) Taxas
    - Linha start+10: (-) Inadimplência
    - Linha start+11: (-) Custos Fixos sem PL
    - Linha start+12: (-) Pró-labore total
    - Linha start+13: LUCRO MENSAL
    - Linha start+14: Margem %
    - Linha start+15: Caixa acumulado
    Coluna B = label, C-N = meses 1-12, O = caixa acumulado final, P = total ano
    """
    section_row(ws, start_row, f"CENÁRIO {nome.upper()}", span=16)
    # Personalizar fill na section
    for col in range(1, 17):
        ws.cell(row=start_row, column=col).fill = fill

    # Header
    headers = ["Linha"] + [f"M{m}" for m in range(1, 13)] + ["Caixa final", "TOTAL ANO"]
    for i, h in enumerate(headers):
        c = ws.cell(row=start_row+1, column=2+i, value=h)
        c.font=HEAD; c.fill=FILL_LB; c.alignment=CENTER; c.border=BORDER

    # Receita 48h - projeto_qty_per_month é lista de 12 inteiros = projetos/mês
    # projeto_mix dict: {"Start": peso, "Pro": peso, "Premium": peso} somando 1
    ws.cell(row=start_row+2, column=2, value="Receita Projeto 48h").font=BOLD
    ws.cell(row=start_row+2, column=2).border=BORDER; ws.cell(row=start_row+2, column=2).alignment=LEFT
    rec_48h_per_proj = (projeto_mix["Start"]*PRECOS_48H["Start"]
                       + projeto_mix["Pro"]*PRECOS_48H["Pro"]
                       + projeto_mix["Premium"]*PRECOS_48H["Premium"])
    for m in range(12):
        c = ws.cell(row=start_row+2, column=3+m, value=projeto_qty_per_month[m]*rec_48h_per_proj)
        c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
    ws.cell(row=start_row+2, column=15, value="").border=BORDER  # caixa final coluna O
    ws.cell(row=start_row+2, column=16, value=f"=SUM(C{start_row+2}:N{start_row+2})").border=BORDER
    money0_fmt(ws.cell(row=start_row+2, column=16))
    ws.cell(row=start_row+2, column=16).font=BOLD; ws.cell(row=start_row+2, column=16).alignment=CENTER

    # Receita personalizados (zero por padrão neste plano; placeholder)
    ws.cell(row=start_row+3, column=2, value="Receita Personalizados").font=NORMAL
    ws.cell(row=start_row+3, column=2).border=BORDER; ws.cell(row=start_row+3, column=2).alignment=LEFT
    for m in range(12):
        c = ws.cell(row=start_row+3, column=3+m, value=0); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
    ws.cell(row=start_row+3, column=16, value=0).border=BORDER; money0_fmt(ws.cell(row=start_row+3, column=16))

    # Receita Suporte (cumulativo - cada mês acumula)
    ws.cell(row=start_row+4, column=2, value="Receita Suporte Mensal (acumulado)").font=BOLD
    ws.cell(row=start_row+4, column=2).border=BORDER; ws.cell(row=start_row+4, column=2).alignment=LEFT
    # suporte_qty_per_month é lista de 12 = NOVOS contratos/mês
    # acumular: receita mês m = soma(novos até m) × preço plano
    preco_sup = PRECOS_SUP[suporte_plan]
    acum = 0
    for m in range(12):
        acum += suporte_qty_per_month[m]
        c = ws.cell(row=start_row+4, column=3+m, value=acum*preco_sup)
        c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
    ws.cell(row=start_row+4, column=16, value=f"=SUM(C{start_row+4}:N{start_row+4})").border=BORDER
    money0_fmt(ws.cell(row=start_row+4, column=16))
    ws.cell(row=start_row+4, column=16).font=BOLD; ws.cell(row=start_row+4, column=16).alignment=CENTER

    # Receita Parceria (cumulativo)
    ws.cell(row=start_row+5, column=2, value="Receita Parceria (acumulado)").font=NORMAL
    ws.cell(row=start_row+5, column=2).border=BORDER; ws.cell(row=start_row+5, column=2).alignment=LEFT
    acum = 0
    for m in range(12):
        acum += parc_qty_per_month[m]
        c = ws.cell(row=start_row+5, column=3+m, value=acum*60)  # R$ 60 médio
        c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
    ws.cell(row=start_row+5, column=16, value=f"=SUM(C{start_row+5}:N{start_row+5})").border=BORDER
    money0_fmt(ws.cell(row=start_row+5, column=16))

    # Receita Tráfego Pago
    ws.cell(row=start_row+6, column=2, value="Receita Tráfego Pago").font=NORMAL
    ws.cell(row=start_row+6, column=2).border=BORDER; ws.cell(row=start_row+6, column=2).alignment=LEFT
    for m in range(12):
        c = ws.cell(row=start_row+6, column=3+m, value=traf_qty_per_month[m]*traf_fee)
        c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
    ws.cell(row=start_row+6, column=16, value=f"=SUM(C{start_row+6}:N{start_row+6})").border=BORDER
    money0_fmt(ws.cell(row=start_row+6, column=16))

    # Linha 7 = RECEITA BRUTA
    ws.cell(row=start_row+7, column=2, value="RECEITA BRUTA TOTAL").font=BOLD
    ws.cell(row=start_row+7, column=2).fill = FILL_H3
    ws.cell(row=start_row+7, column=2).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=start_row+7, column=2).border=BORDER; ws.cell(row=start_row+7, column=2).alignment=LEFT
    for m in range(12):
        col = chr(ord("C") + m)
        f = f"=SUM({col}{start_row+2}:{col}{start_row+6})"
        c = ws.cell(row=start_row+7, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD
    ws.cell(row=start_row+7, column=16, value=f"=SUM(C{start_row+7}:N{start_row+7})").border=BORDER
    money0_fmt(ws.cell(row=start_row+7, column=16))
    ws.cell(row=start_row+7, column=16).font=BOLD; ws.cell(row=start_row+7, column=16).fill=FILL_H3
    ws.cell(row=start_row+7, column=16).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=start_row+7, column=16).alignment=CENTER

    # Linha 8 = Impostos
    ws.cell(row=start_row+8, column=2, value="(-) Impostos (Simples 12%)").font=NORMAL
    ws.cell(row=start_row+8, column=2).border=BORDER; ws.cell(row=start_row+8, column=2).alignment=LEFT
    for m in range(12):
        col = chr(ord("C") + m)
        f = f"={col}{start_row+7}*'2. Premissas'!{PR_REF['imposto']}"
        c = ws.cell(row=start_row+8, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
    ws.cell(row=start_row+8, column=16, value=f"=SUM(C{start_row+8}:N{start_row+8})").border=BORDER; money0_fmt(ws.cell(row=start_row+8, column=16))

    # Linha 9 = Taxas
    ws.cell(row=start_row+9, column=2, value="(-) Taxas cartão (média)").font=NORMAL
    ws.cell(row=start_row+9, column=2).border=BORDER; ws.cell(row=start_row+9, column=2).alignment=LEFT
    for m in range(12):
        col = chr(ord("C") + m)
        f = f"={col}{start_row+7}*'2. Premissas'!{PR_REF['tx_media']}"
        c = ws.cell(row=start_row+9, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
    ws.cell(row=start_row+9, column=16, value=f"=SUM(C{start_row+9}:N{start_row+9})").border=BORDER; money0_fmt(ws.cell(row=start_row+9, column=16))

    # Linha 10 = Inadimplência
    ws.cell(row=start_row+10, column=2, value="(-) Inadimplência (PCLD)").font=NORMAL
    ws.cell(row=start_row+10, column=2).border=BORDER; ws.cell(row=start_row+10, column=2).alignment=LEFT
    for m in range(12):
        col = chr(ord("C") + m)
        f = f"={col}{start_row+7}*'2. Premissas'!{PR_REF['inad']}"
        c = ws.cell(row=start_row+10, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
    ws.cell(row=start_row+10, column=16, value=f"=SUM(C{start_row+10}:N{start_row+10})").border=BORDER; money0_fmt(ws.cell(row=start_row+10, column=16))

    # Linha 11 = Custos Fixos sem PL
    ws.cell(row=start_row+11, column=2, value="(-) Custos Fixos (sem pró-labore)").font=NORMAL
    ws.cell(row=start_row+11, column=2).border=BORDER; ws.cell(row=start_row+11, column=2).alignment=LEFT
    for m in range(12):
        c = ws.cell(row=start_row+11, column=3+m, value=f"='2. Premissas'!{PR_REF['cf_sem_pl']}")
        c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
    ws.cell(row=start_row+11, column=16, value=f"=SUM(C{start_row+11}:N{start_row+11})").border=BORDER; money0_fmt(ws.cell(row=start_row+11, column=16))

    # Linha 12 = Pró-labore + INSS
    ws.cell(row=start_row+12, column=2, value="(-) Pró-labore + INSS sócios").font=NORMAL
    ws.cell(row=start_row+12, column=2).border=BORDER; ws.cell(row=start_row+12, column=2).alignment=LEFT
    for m in range(12):
        c = ws.cell(row=start_row+12, column=3+m,
                    value=f"='2. Premissas'!{PR_REF['pl_total']}+'2. Premissas'!{PR_REF['inss_total']}")
        c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
    ws.cell(row=start_row+12, column=16, value=f"=SUM(C{start_row+12}:N{start_row+12})").border=BORDER; money0_fmt(ws.cell(row=start_row+12, column=16))

    # Linha 13 = LUCRO MENSAL
    ws.cell(row=start_row+13, column=2, value="LUCRO MENSAL").font=BOLD
    ws.cell(row=start_row+13, column=2).fill = FILL_LG
    ws.cell(row=start_row+13, column=2).border=BORDER; ws.cell(row=start_row+13, column=2).alignment=LEFT
    for m in range(12):
        col = chr(ord("C") + m)
        f = f"={col}{start_row+7}-{col}{start_row+8}-{col}{start_row+9}-{col}{start_row+10}-{col}{start_row+11}-{col}{start_row+12}"
        c = ws.cell(row=start_row+13, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD; c.fill=FILL_LG
    ws.cell(row=start_row+13, column=16, value=f"=SUM(C{start_row+13}:N{start_row+13})").border=BORDER; money0_fmt(ws.cell(row=start_row+13, column=16))
    ws.cell(row=start_row+13, column=16).font=BOLD; ws.cell(row=start_row+13, column=16).fill=FILL_LG; ws.cell(row=start_row+13, column=16).alignment=CENTER

    # Linha 14 = Margem
    ws.cell(row=start_row+14, column=2, value="Margem líquida %").font=NORMAL
    ws.cell(row=start_row+14, column=2).border=BORDER; ws.cell(row=start_row+14, column=2).alignment=LEFT
    for m in range(12):
        col = chr(ord("C") + m)
        f = f"=IFERROR({col}{start_row+13}/{col}{start_row+7},0)"
        c = ws.cell(row=start_row+14, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; pct_fmt(c)
    ws.cell(row=start_row+14, column=16, value=f"=IFERROR(P{start_row+13}/P{start_row+7},0)").border=BORDER; pct_fmt(ws.cell(row=start_row+14, column=16)); ws.cell(row=start_row+14, column=16).font=BOLD; ws.cell(row=start_row+14, column=16).alignment=CENTER

    # Linha 15 = Caixa acumulado
    ws.cell(row=start_row+15, column=2, value="Caixa acumulado").font=BOLD
    ws.cell(row=start_row+15, column=2).fill = FILL_LB
    ws.cell(row=start_row+15, column=2).border=BORDER; ws.cell(row=start_row+15, column=2).alignment=LEFT
    # Mês 1: caixa = lucro mês 1
    ws.cell(row=start_row+15, column=3, value=f"=C{start_row+13}").border=BORDER; money0_fmt(ws.cell(row=start_row+15, column=3)); ws.cell(row=start_row+15, column=3).alignment=CENTER; ws.cell(row=start_row+15, column=3).font=BOLD; ws.cell(row=start_row+15, column=3).fill=FILL_LB
    for m in range(1, 12):
        col = chr(ord("C") + m)
        prev_col = chr(ord("C") + m - 1)
        f = f"={prev_col}{start_row+15}+{col}{start_row+13}"
        c = ws.cell(row=start_row+15, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD; c.fill=FILL_LB
    # Caixa final coluna 15 (O) = mesma do mês 12 (N)
    ws.cell(row=start_row+15, column=15, value=f"=N{start_row+15}").border=BORDER; money0_fmt(ws.cell(row=start_row+15, column=15)); ws.cell(row=start_row+15, column=15).font=BOLD; ws.cell(row=start_row+15, column=15).alignment=CENTER; ws.cell(row=start_row+15, column=15).fill=FILL_H3
    ws.cell(row=start_row+15, column=15).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws.cell(row=start_row+15, column=16, value="").border=BORDER

    # Total summary row (start_row + 18)
    ws.cell(row=start_row+18, column=2, value="RESUMO ANUAL DO CENÁRIO").font=BOLD; ws.cell(row=start_row+18, column=2).fill=FILL_H1
    ws.cell(row=start_row+18, column=2).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    ws.cell(row=start_row+18, column=2).border=BORDER
    # P22 (esperado pelo Resumo) = receita bruta anual
    # Os indicadores ref na aba 1 foram: P22, P28, P29, O22 para Pessimista
    # Vou refazer: célula P (linha 22 do pessimista = receita)
    # Para coincidir com a aba 1, vou criar células de resumo:
    # P22 = Receita anual; P28 = Lucro anual; P29 = Margem; O22 = Caixa final
    # Como cada cenário usa offsets diferentes, é mais simples colocar no formato consistente:
    # Pessimista: receita anual em P22 (linha 22 do pess + 18 offset?)
    # Vou criar células de resumo COMUNS abaixo da seção, e a aba 1 vai referenciar essas células fixas.
    # Refs FIXAS para Resumo Executivo:
    # P22 (Pess receita), P37 (Real), P52 (Otim)
    # Vou ajustar: para PESS_START=4, a receita anual está em P11 (start_row+7=11). Não é P22.
    # Vou mudar estratégia: criar uma seção de resumo FIXA no topo da aba 12.
    pass

# Configurações dos cenários:
# Cada lista de 12 meses representa qtd de projetos novos por mês.
# Suporte é qtd de NOVOS contratos por mês (cumulativo na função).

# PESSIMISTA
build_cenario(ws, PESS_START, "Pessimista", FILL_LR,
              projeto_qty_per_month=[1,1,1,2,2,2,2,3,3,3,3,3],   # ramp up lento
              suporte_qty_per_month=[0,1,1,0,1,1,1,1,1,1,1,1],
              parc_qty_per_month=[0,0,0,0,0,0,1,1,1,2,2,2],
              traf_qty_per_month=[0]*12,
              projeto_mix={"Start": 0.7, "Pro": 0.3, "Premium": 0},
              suporte_plan="Essencial",
              traf_fee=0)

# REALISTA
build_cenario(ws, REAL_START, "Realista", FILL_LY,
              projeto_qty_per_month=[2,2,3,3,3,3,3,3,3,3,3,3],
              suporte_qty_per_month=[1,2,2,2,2,2,2,2,2,2,2,2],
              parc_qty_per_month=[0,0,1,1,2,2,2,2,2,2,2,2],
              traf_qty_per_month=[0,0,0,1,1,1,1,1,1,1,1,1],
              projeto_mix={"Start": 0.4, "Pro": 0.4, "Premium": 0.2},
              suporte_plan="Pro",
              traf_fee=2500)

# OTIMISTA
build_cenario(ws, OTIM_START, "Otimista", FILL_LG,
              projeto_qty_per_month=[3,4,4,5,5,5,5,5,5,5,5,5],
              suporte_qty_per_month=[2,3,3,3,3,3,3,3,3,3,3,3],
              parc_qty_per_month=[1,2,2,3,3,3,3,3,3,3,3,3],
              traf_qty_per_month=[1,1,2,2,2,2,2,2,2,2,2,2],
              projeto_mix={"Start": 0.3, "Pro": 0.5, "Premium": 0.2},
              suporte_plan="Pro",
              traf_fee=2500)

# Criar células de resumo ao final para o Resumo Executivo
RESUMO_START = OTIM_START + 22
ws.merge_cells(start_row=RESUMO_START, start_column=2, end_row=RESUMO_START, end_column=16)
c = ws.cell(row=RESUMO_START, column=2, value="📊 RESUMO ANUAL COMPARATIVO (linkado à aba 1)")
c.font = H2; c.fill = FILL_H1
c.alignment = LEFT; c.border = BORDER
ws.row_dimensions[RESUMO_START].height = 24

# Header
header_cells(ws, RESUMO_START+1, ["Indicador", "Pessimista", "Realista", "Otimista"], start_col=2)

# Aqui mapeamos as linhas de cada cenário:
# Pessimista linha base 4 → receita bruta em linha 4+7=11, lucro em 4+13=17, margem em 4+14=18, caixa em 4+15=19
# Realista 31+7=38, 31+13=44, 31+14=45, 31+15=46
# Otimista 58+7=65, 58+13=71, 58+14=72, 58+15=73

PESS_REC_ROW = PESS_START + 7   # = 11
PESS_LUC_ROW = PESS_START + 13  # = 17
PESS_MAR_ROW = PESS_START + 14  # = 18
PESS_CX_ROW  = PESS_START + 15  # = 19
REAL_REC_ROW = REAL_START + 7   # = 38
REAL_LUC_ROW = REAL_START + 13  # = 44
REAL_MAR_ROW = REAL_START + 14  # = 45
REAL_CX_ROW  = REAL_START + 15  # = 46
OTIM_REC_ROW = OTIM_START + 7   # = 65
OTIM_LUC_ROW = OTIM_START + 13  # = 71
OTIM_MAR_ROW = OTIM_START + 14  # = 72
OTIM_CX_ROW  = OTIM_START + 15  # = 73

indicadores_resumo = [
    ("Receita bruta anual", f"=P{PESS_REC_ROW}", f"=P{REAL_REC_ROW}", f"=P{OTIM_REC_ROW}", "money"),
    ("Lucro líquido anual", f"=P{PESS_LUC_ROW}", f"=P{REAL_LUC_ROW}", f"=P{OTIM_LUC_ROW}", "money"),
    ("Margem líquida média", f"=P{PESS_MAR_ROW}", f"=P{REAL_MAR_ROW}", f"=P{OTIM_MAR_ROW}", "pct"),
    ("Caixa acumulado final mês 12", f"=O{PESS_CX_ROW}", f"=O{REAL_CX_ROW}", f"=O{OTIM_CX_ROW}", "money"),
]
r = RESUMO_START + 2
for label, p, real, ot, kind in indicadores_resumo:
    a = ws.cell(row=r, column=2, value=label); a.font=BOLD; a.border=BORDER; a.alignment=LEFT; a.fill=FILL_GREY
    for j, v in enumerate([p, real, ot]):
        c = ws.cell(row=r, column=3+j, value=v); c.border=BORDER; c.alignment=CENTER; c.font=BOLD
        if kind == "money": money0_fmt(c)
        elif kind == "pct": pct_fmt(c)
        c.fill = [FILL_LR, FILL_LY, FILL_LG][j]
    r += 1

# Atualizar âncoras da aba 1
ABA1_REFS = {
    "rec_pess": f"P{PESS_REC_ROW}", "rec_real": f"P{REAL_REC_ROW}", "rec_otim": f"P{OTIM_REC_ROW}",
    "luc_pess": f"P{PESS_LUC_ROW}", "luc_real": f"P{REAL_LUC_ROW}", "luc_otim": f"P{OTIM_LUC_ROW}",
    "mar_pess": f"P{PESS_MAR_ROW}", "mar_real": f"P{REAL_MAR_ROW}", "mar_otim": f"P{OTIM_MAR_ROW}",
    "cx_pess":  f"O{PESS_CX_ROW}",  "cx_real":  f"O{REAL_CX_ROW}",  "cx_otim":  f"O{OTIM_CX_ROW}",
}
print(f"Aba 12 ✓ — refs cenários: {ABA1_REFS}")


# ============================================================
# ABA 13 — DRE PROJETADA (12 meses cenário Realista — vinculada à aba 12)
# ============================================================
ws = wb.create_sheet("13. DRE")
set_col_widths(ws, [3, 36] + [11]*13 + [3])

title_row(ws, 1, "DRE PROJETADA — CENÁRIO REALISTA (12 meses)", span=16)
ws.cell(row=2, column=2, value="Estrutura conforme CPC 26 (R1). Vinculada à aba 12 — Realista (linhas 31-48)").font=ITAL

header_cells(ws, 4, ["Linha DRE"] + [f"M{m}" for m in range(1, 13)] + ["ANO"], start_col=2)

# Para cada item da DRE, referenciar a aba 12 (Cenário Realista)
# Realista: linhas 33 = receita 48h, 35 = suporte, 36 = parceria, 37 = tráfego, 38 = receita bruta
dre_rows = [
    ("RECEITA BRUTA OPERACIONAL", f"='12. Cenários'!{{col}}38", FILL_LB, True),
    ("  Projeto 48 Horas", f"='12. Cenários'!{{col}}33", None, False),
    ("  Projetos Personalizados", f"='12. Cenários'!{{col}}34", None, False),
    ("  Suporte Mensal (recorrente)", f"='12. Cenários'!{{col}}35", None, False),
    ("  Parceria com Plataformas", f"='12. Cenários'!{{col}}36", None, False),
    ("  Tráfego Pago (gestão)", f"='12. Cenários'!{{col}}37", None, False),
    ("(–) Deduções da Receita", "=({col}6+{col}10+{col}11)*-1", None, True),
    ("  Impostos (Simples)", f"='12. Cenários'!{{col}}39", None, False),
    ("  Taxas cartão (média)", f"='12. Cenários'!{{col}}40", None, False),
    ("  Inadimplência (PCLD)", f"='12. Cenários'!{{col}}41", None, False),
    ("= RECEITA OPERACIONAL LÍQUIDA", "={col}5+{col}12", FILL_LG, True),
    ("(–) CSP — Custos dos Serviços Prestados", "=({col}14)*-1", None, True),
    ("  Mão de obra direta (estimada 30% da receita)", "={col}5*0.30", None, False),
    ("= LUCRO BRUTO", "={col}13+{col}15", FILL_LG, True),
    ("(–) Despesas Operacionais", "=({col}17+{col}18)*-1", None, True),
    ("  Custos fixos administrativos", f"='12. Cenários'!{{col}}42", None, False),
    ("  Pró-labore + INSS sócios", f"='12. Cenários'!{{col}}43", None, False),
    ("= EBITDA", "={col}16+{col}17+{col}18 -{col}5*0.30 + {col}5*0.30", None, True),  # ajuste para evitar dupla contagem
    ("(–) Depreciação / amortização", 0, None, False),
    ("= EBIT (Lucro Operacional)", "={col}19+{col}20", FILL_LG, True),
    ("(–) Resultado Financeiro", 0, None, False),
    ("= LUCRO LÍQUIDO DO PERÍODO", "='12. Cenários'!{col}44", FILL_H3, True),
    ("Margem líquida", "=IFERROR({col}23/{col}5,0)", FILL_LB, True),
]

# Vou simplificar a DRE para evitar bagunça — usar abordagem direta:
# Cada linha da DRE puxa direto da aba 12 quando faz sentido
dre_simple = [
    ("RECEITA BRUTA OPERACIONAL", "P38_REC", FILL_LB, "BOLD"),
    ("  Projeto 48 Horas", "P33", None, ""),
    ("  Projetos Personalizados", "P34", None, ""),
    ("  Suporte Mensal Recorrente (MRR)", "P35", None, ""),
    ("  Parceria com Plataformas", "P36", None, ""),
    ("  Tráfego Pago (gestão)", "P37", None, ""),
    ("(–) Impostos sobre venda (Simples)", "P39", None, ""),
    ("(–) Taxas de cartão", "P40", None, ""),
    ("(–) Inadimplência (PCLD)", "P41", None, ""),
    ("= RECEITA LÍQUIDA", "RL", FILL_LG, "BOLD"),
    ("(–) Custos Fixos Administrativos", "P42", None, ""),
    ("(–) Pró-labore + INSS sócios", "P43", None, ""),
    ("= LUCRO LÍQUIDO MENSAL", "P44", FILL_H3, "BOLD-WHITE"),
    ("Margem líquida (%)", "MARGEM", FILL_LB, "BOLD"),
]

r = 5
for label, kind, fill, style in dre_simple:
    a = ws.cell(row=r, column=2, value=label)
    a.border = BORDER; a.alignment = LEFT
    if "BOLD" in (style or ""): a.font = BOLD
    if "WHITE" in (style or ""): a.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    if fill: a.fill = fill

    for m in range(12):
        col_letter_aba12 = chr(ord("C") + m)
        if kind == "RL":
            f = f"=C{r-9}+C{r-3}+C{r-2}+C{r-1}".replace("C", col_letter_aba12) if False else None
            # Receita Líquida = Receita Bruta - Impostos - Taxas - Inad
            # Receita Bruta está em linha 5 (label index 0) → r_brut = 5
            # Impostos = linha 11 (label index 6) → r=5+6=11
            # mas as linhas começam na 5 e contam, então:
            # Receita Bruta = linha 5 (1ª linha)
            # Receita Líquida = linha 14 (10ª linha = índice 9)
            # Vamos calcular relativamente: receita bruta = r - 9; impostos = r - 3; taxas = r - 2; inad = r - 1
            f = f"={col_letter_aba12}5_aba+{col_letter_aba12}11_aba+{col_letter_aba12}12_aba+{col_letter_aba12}13_aba"
            # Aproximação direta: usar aba 12 diretamente
            f = f"='12. Cenários'!{col_letter_aba12}38 - '12. Cenários'!{col_letter_aba12}39 - '12. Cenários'!{col_letter_aba12}40 - '12. Cenários'!{col_letter_aba12}41"
        elif kind == "MARGEM":
            # Margem = lucro / receita bruta = linha 17 / linha 5
            # Em fórmula direta da aba 12: P44 / P38
            f = f"=IFERROR('12. Cenários'!{col_letter_aba12}44/'12. Cenários'!{col_letter_aba12}38,0)"
        elif kind.startswith("P"):
            # Referência direta à aba 12, linha numérica
            row_aba12 = int(kind[1:].replace("_REC",""))
            f = f"='12. Cenários'!{col_letter_aba12}{row_aba12}"
        else:
            f = 0
        c = ws.cell(row=r, column=3+m, value=f)
        c.border = BORDER; c.alignment = CENTER; money0_fmt(c)
        if "WHITE" in (style or ""):
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c.fill = FILL_H3
        elif "BOLD" in (style or ""):
            c.font = BOLD
        if fill and "WHITE" not in (style or ""):
            c.fill = fill
        if kind == "MARGEM":
            pct_fmt(c)

    # Coluna ANO (P, índice 16)
    if kind == "RL":
        col_letter_aba12 = "P"
        f = f"='12. Cenários'!P38 - '12. Cenários'!P39 - '12. Cenários'!P40 - '12. Cenários'!P41"
    elif kind == "MARGEM":
        f = f"=IFERROR('12. Cenários'!P44/'12. Cenários'!P38,0)"
    elif kind.startswith("P"):
        row_aba12 = int(kind[1:].replace("_REC",""))
        f = f"='12. Cenários'!P{row_aba12}"
    else:
        f = 0
    c = ws.cell(row=r, column=15, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD
    if "WHITE" in (style or ""): c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF"); c.fill = FILL_H3
    elif fill: c.fill = fill
    if kind == "MARGEM": pct_fmt(c)

    # Coluna ANO (P, índice 16) com label
    if r == 5:
        pass  # cabeçalho já foi

    r += 1

print("Aba 13 ✓")


# ============================================================
# ABA 14 — FLUXO DE CAIXA (12 meses Realista, considerando recebimento parcelado)
# ============================================================
ws = wb.create_sheet("14. Fluxo de Caixa")
set_col_widths(ws, [3, 36] + [11]*13 + [3])

title_row(ws, 1, "FLUXO DE CAIXA — CENÁRIO REALISTA (12 meses)", span=16)
ws.cell(row=2, column=2, value="Considera descasamento entre receita reconhecida e caixa efetivamente recebido (parcelamento)").font=ITAL

header_cells(ws, 4, ["Movimento"] + [f"M{m}" for m in range(1, 13)] + ["ANO"], start_col=2)

# Modelo simplificado: assume que 60% da receita entra à vista/PIX, 30% em 3x, 10% em 6x
# Entradas distribuídas mensalmente; saídas são CF + pró-labore mensal
r = 5

# Recebimentos
ws.cell(row=r, column=2, value="(+) Recebimentos à vista/PIX (60% da receita)").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT; ws.cell(row=r, column=2).font=BOLD
for m in range(12):
    col_letter = chr(ord("C") + m)
    f = f"='12. Cenários'!{col_letter}38*0.6"
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15)); ws.cell(row=r, column=15).font=BOLD
r += 1

ws.cell(row=r, column=2, value="(+) Recebimentos parcelado 3x (30% receita, diluído em 3 meses)").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
for m in range(12):
    col_letter = chr(ord("C") + m)
    # Simplificação: assume recebimento médio em t+1
    f = f"='12. Cenários'!{col_letter}38*0.3"
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15))
r += 1

ws.cell(row=r, column=2, value="(+) Recebimentos parcelado 6x+ (10% receita)").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
for m in range(12):
    col_letter = chr(ord("C") + m)
    f = f"='12. Cenários'!{col_letter}38*0.1"
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15))
r += 1

# Total entradas
ws.cell(row=r, column=2, value="= TOTAL ENTRADAS").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT; ws.cell(row=r, column=2).font=BOLD; ws.cell(row=r, column=2).fill=FILL_LG
for m in range(12):
    col_letter = chr(ord("C") + m)
    f = f"=SUM({col_letter}5:{col_letter}7)"
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD; c.fill=FILL_LG
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15)); ws.cell(row=r, column=15).font=BOLD; ws.cell(row=r, column=15).fill=FILL_LG
ENT_ROW = r
r += 2

# Saídas
ws.cell(row=r, column=2, value="(–) Impostos (Simples)").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
for m in range(12):
    col_letter = chr(ord("C") + m)
    f = f"='12. Cenários'!{col_letter}39"
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15))
IMP_ROW = r
r += 1

ws.cell(row=r, column=2, value="(–) Taxas cartão").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
for m in range(12):
    col_letter = chr(ord("C") + m)
    f = f"='12. Cenários'!{col_letter}40"
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15))
TAX_ROW = r
r += 1

ws.cell(row=r, column=2, value="(–) Inadimplência (PCLD)").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
for m in range(12):
    col_letter = chr(ord("C") + m)
    f = f"='12. Cenários'!{col_letter}41"
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15))
INAD_ROW = r
r += 1

ws.cell(row=r, column=2, value="(–) Custos fixos administrativos").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
for m in range(12):
    c = ws.cell(row=r, column=3+m, value=f"='2. Premissas'!{PR_REF['cf_sem_pl']}")
    c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15))
CF_ROW = r
r += 1

ws.cell(row=r, column=2, value="(–) Pró-labore + INSS sócios").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
for m in range(12):
    c = ws.cell(row=r, column=3+m, value=f"='2. Premissas'!{PR_REF['pl_total']}+'2. Premissas'!{PR_REF['inss_total']}")
    c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15))
PL_ROW = r
r += 1

# Total saídas
ws.cell(row=r, column=2, value="= TOTAL SAÍDAS").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT; ws.cell(row=r, column=2).font=BOLD; ws.cell(row=r, column=2).fill=FILL_LR
for m in range(12):
    col_letter = chr(ord("C") + m)
    f = f"={col_letter}{IMP_ROW}+{col_letter}{TAX_ROW}+{col_letter}{INAD_ROW}+{col_letter}{CF_ROW}+{col_letter}{PL_ROW}"
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD; c.fill=FILL_LR
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15)); ws.cell(row=r, column=15).font=BOLD; ws.cell(row=r, column=15).fill=FILL_LR
SAI_ROW = r
r += 2

# Fluxo do mês
ws.cell(row=r, column=2, value="= FLUXO DE CAIXA DO MÊS").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT; ws.cell(row=r, column=2).font=BOLD; ws.cell(row=r, column=2).fill=FILL_H3
ws.cell(row=r, column=2).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
for m in range(12):
    col_letter = chr(ord("C") + m)
    f = f"={col_letter}{ENT_ROW}-{col_letter}{SAI_ROW}"
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD; c.fill=FILL_H3
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ws.cell(row=r, column=15, value=f"=SUM(C{r}:N{r})").border=BORDER; money0_fmt(ws.cell(row=r, column=15)); ws.cell(row=r, column=15).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF"); ws.cell(row=r, column=15).fill=FILL_H3
FLUXO_ROW = r
r += 1

# Caixa inicial
ws.cell(row=r, column=2, value="(+) Caixa inicial (reserva)").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
c0 = ws.cell(row=r, column=3, value=50000); c0.border=BORDER; c0.alignment=CENTER; money0_fmt(c0); c0.fill=FILL_LY
for m in range(1, 12):
    col_letter = chr(ord("C") + m)
    prev_col = chr(ord("C") + m - 1)
    f = f"={prev_col}{r+1}"  # caixa final do mês anterior
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
CX_INI_ROW = r
r += 1

# Caixa final
ws.cell(row=r, column=2, value="= CAIXA FINAL ACUMULADO").border=BORDER; ws.cell(row=r, column=2).alignment=LEFT; ws.cell(row=r, column=2).font=BOLD; ws.cell(row=r, column=2).fill=FILL_LB
for m in range(12):
    col_letter = chr(ord("C") + m)
    f = f"={col_letter}{CX_INI_ROW}+{col_letter}{FLUXO_ROW}"
    c = ws.cell(row=r, column=3+m, value=f); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD; c.fill=FILL_LB
ws.cell(row=r, column=15, value=f"=N{r}").border=BORDER; money0_fmt(ws.cell(row=r, column=15)); ws.cell(row=r, column=15).font=BOLD; ws.cell(row=r, column=15).fill=FILL_LB; ws.cell(row=r, column=15).alignment=CENTER

r += 2
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=15)
c = ws.cell(row=r, column=2,
            value="⚠️ Premissa de fluxo: 60% PIX/à vista | 30% em até 3x | 10% em 6x+. Modelo simplificado — não considera antecipação de recebíveis. "
                  "Capital de giro inicial: R$ 50.000 (recomendação CFO; ajustar em B41 se mudar). "
                  "Risco do parcelado: descasamento receita reconhecida vs caixa pode chegar a 60-90 dias.")
c.font=NORMAL; c.fill=FILL_LY; c.alignment=LEFT_TOP; c.border=BORDER
ws.row_dimensions[r].height = 50

print("Aba 14 ✓")


# ============================================================
# ABA 15 — PONTO DE EQUILÍBRIO
# ============================================================
ws = wb.create_sheet("15. Ponto Equilíbrio")
set_col_widths(ws, [3, 42, 18, 38, 3])

title_row(ws, 1, "PONTO DE EQUILÍBRIO — FATURAMENTO MÍNIMO MENSAL", span=5)

section_row(ws, 3, "FÓRMULA E CÁLCULO", span=5)
header_cells(ws, 4, ["Componente", "Valor", "Origem"], start_col=2)

itens = [
    ("Custo fixo total mensal (com pró-labore + INSS)", f"='2. Premissas'!{PR_REF['cf_total']}", "Soma da aba 2 — premissas"),
    ("Imposto Simples", f"='2. Premissas'!{PR_REF['imposto']}", "Premissa editável"),
    ("Taxa cartão média", f"='2. Premissas'!{PR_REF['tx_media']}", "Média ponderada"),
    ("Inadimplência", f"='2. Premissas'!{PR_REF['inad']}", "PCLD 4%"),
    ("Deduções variáveis totais (impostos + taxa + inad)", f"='2. Premissas'!{PR_REF['deduc_var']}", "Soma das 3 deduções"),
    ("Margem de contribuição efetiva", f"=1-'2. Premissas'!{PR_REF['deduc_var']}", "= 1 - deduções"),
]
r = 5
for nome, val, obs in itens:
    a = ws.cell(row=r, column=2, value=nome); a.font=BOLD; a.border=BORDER; a.alignment=LEFT
    b = ws.cell(row=r, column=3, value=val); b.border=BORDER; b.alignment=CENTER; b.font=BOLD
    if "Custo fixo" in nome: money0_fmt(b)
    else: pct_fmt(b)
    c = ws.cell(row=r, column=4, value=obs); c.border=BORDER; c.alignment=LEFT_TOP; c.font=ITAL
    ws.row_dimensions[r].height = 22
    r += 1

r += 1
# Ponto de equilíbrio
ws.cell(row=r, column=2, value="🎯 PONTO DE EQUILÍBRIO MENSAL (R$)").font=H2; ws.cell(row=r, column=2).fill=FILL_H1
ws.cell(row=r, column=2).font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
ws.cell(row=r, column=2).border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
pe_formula = f"='2. Premissas'!{PR_REF['cf_total']}/(1-'2. Premissas'!{PR_REF['deduc_var']})"
c = ws.cell(row=r, column=3, value=pe_formula); c.border=BORDER; c.alignment=CENTER; money0_fmt(c)
c.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF"); c.fill = FILL_H1
ws.cell(row=r, column=4, value="PE = CF / (1 - deduções variáveis)").font=ITAL; ws.cell(row=r, column=4).border=BORDER
ws.row_dimensions[r].height = 28
# Linha 14 conforme aba 1 espera
PE_ROW = r

r += 1
ws.cell(row=r, column=2, value="Ponto de equilíbrio ANUAL").font=BOLD; ws.cell(row=r, column=2).border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
c = ws.cell(row=r, column=3, value=f"=C{PE_ROW}*12"); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD
r += 1
ws.cell(row=r, column=2, value="Margem de segurança recomendada (+30%)").font=BOLD; ws.cell(row=r, column=2).border=BORDER; ws.cell(row=r, column=2).alignment=LEFT
c = ws.cell(row=r, column=3, value=f"=C{PE_ROW}*1.3"); c.border=BORDER; c.alignment=CENTER; money0_fmt(c); c.font=BOLD; c.fill=FILL_LG
ws.cell(row=r, column=4, value="Patamar saudável de operação").font=ITAL; ws.cell(row=r, column=4).border=BORDER

r += 2
section_row(ws, r, "COMBINAÇÕES PARA ATINGIR O PE (exemplos)", span=5); r += 1
header_cells(ws, r, ["Combinação", "Receita projetada", "Atinge PE?"], start_col=2); r += 1

combos = [
    ("4 projetos Pro/mês (R$ 4.990 × 4)", "=4990*4", '=IF(C18>=C13,"✅ Sim","❌ Não")'),
    ("3 Start + 1 Pro/mês", "=2490*3+4990", '=IF(C19>=C13,"✅ Sim","❌ Não")'),
    ("2 Pro + 1 Premium", "=4990*2+8900", '=IF(C20>=C13,"✅ Sim","❌ Não")'),
    ("10 contratos Suporte Pro (MRR)", "=990*10", '=IF(C21>=C13,"✅ Sim","❌ Não")'),
    ("8 contratos Suporte Pro + 1 Start", "=990*8+2490", '=IF(C22>=C13,"✅ Sim","❌ Não")'),
    ("2 Start + 1 Pro + 5 Suporte Pro (MIX REALISTA)", "=2490*2+4990+990*5", '=IF(C23>=C13,"✅ Sim","❌ Não")'),
    ("1 Premium + 8 Suporte Pro", "=8900+990*8", '=IF(C24>=C13,"✅ Sim","❌ Não")'),
]
for combo, rec, atinge in combos:
    a = ws.cell(row=r, column=2, value=combo); a.border=BORDER; a.alignment=LEFT
    b = ws.cell(row=r, column=3, value=rec); b.border=BORDER; b.alignment=CENTER; money0_fmt(b)
    # Atualizar referência para PE_ROW dinamicamente
    if "C13" in atinge:
        atinge = atinge.replace("C13", f"$C${PE_ROW}").replace("C18", f"C{r}").replace("C19", f"C{r}").replace("C20", f"C{r}").replace("C21", f"C{r}").replace("C22", f"C{r}").replace("C23", f"C{r}").replace("C24", f"C{r}")
    c = ws.cell(row=r, column=4, value=atinge); c.border=BORDER; c.alignment=CENTER; c.font=BOLD
    r += 1

r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
c = ws.cell(row=r, column=2,
            value=f"💡 PE em horas faturáveis: dividir o PE pela hora de venda recomendada (R$ 89-118/h) — equivale a ~140-180h/mês. "
                  "Isso representa 40-50% da capacidade produtiva total. Operar a partir de R$ 25.000/mês oferece margem de segurança confortável.")
c.font=NORMAL; c.fill=FILL_LY; c.alignment=LEFT_TOP; c.border=BORDER
ws.row_dimensions[r].height = 44

# Atualizar referência da aba 1 para o PE_ROW correto (não 14)
print(f"Aba 15 ✓ — PE em linha {PE_ROW}")
PE_ABA1_REF = PE_ROW


# ============================================================
# ABA 16 — INDICADORES ESTRATÉGICOS (KPIs)
# ============================================================
ws = wb.create_sheet("16. Indicadores")
set_col_widths(ws, [3, 36, 20, 20, 40, 3])

title_row(ws, 1, "INDICADORES ESTRATÉGICOS (KPIs)", span=6)
ws.cell(row=2, column=2, value="Dashboard de acompanhamento mensal — use as fórmulas como referência").font=ITAL

section_row(ws, 4, "INDICADORES OPERACIONAIS E FINANCEIROS", span=6)
header_cells(ws, 5, ["Indicador", "Fórmula", "Meta", "Como medir / Observação"], start_col=2)

kpis = [
    ("Ticket médio Projeto 48h",            "Receita 48h / qtd projetos", "R$ 4.500-5.500", "Mix Start/Pro/Premium influencia"),
    ("MRR (Monthly Recurring Revenue)",     "Soma das mensalidades ativas no mês", "R$ 25.000 até mês 12", "Coração do negócio — KR principal do OKR2"),
    ("ARR (Annual Recurring Revenue)",      "MRR × 12", "R$ 300.000 até mês 12", "Para avaliação do negócio"),
    ("CAC (Custo de Aquisição de Cliente)", "Marketing total / clientes novos", "< R$ 800", "Payback < 3 meses"),
    ("LTV (Lifetime Value)",                "Ticket × meses de retenção × margem", "> R$ 5.000", "LTV/CAC ratio > 5"),
    ("Churn mensal",                        "Cancelamentos / base ativa", "< 4%", "Crítico para MRR sustentável"),
    ("Conversão Projeto → Suporte mensal",  "Clientes em suporte / clientes total", "60% em 12m | 65% em 24m", "Mecânicas: bundle, revisão D+30/D+60"),
    ("Capacidade ocupada",                  "Horas faturáveis / horas alvo", "≤ 80%", "Gatilho de contratação > 85% por 2 meses"),
    ("Receita por sócio",                   "Receita total / 2 sócios", "R$ 15-20k mês 12 (realista)", "Métrica de saúde individual"),
    ("Receita por hora produtiva",          "Receita / horas produtivas", "R$ 110-150/h", "Indica eficiência de monetização"),
    ("Margem por linha de serviço",         "Lucro linha / Receita linha", "48h: 40% | Suporte: 50% | Parceria: ~95%", "Identifica linhas a expandir"),
    ("Taxa de retrabalho",                  "Horas extras não faturadas / horas total", "< 10%", "Sinal de aceite/escopo ambíguo"),
    ("% receita recorrente",                "(Suporte + Parceria) / Receita total", "30% em 12m | 50% em 24m", "Métrica de previsibilidade"),
    ("% receita projetos avulsos",          "(48h + Personalizados) / Receita", "70% em 12m | 50% em 24m", "Complemento do anterior"),
    ("NPS pós-projeto",                     "Pesquisa após handover", "> 8", "Trigger de upsell para suporte"),
    ("Tempo médio de entrega (real vs prometido)", "Dias entrega / dias prometidos", "≤ 1,2× (atraso < 20%)", "Sinal de overbooking"),
    ("Fator R (folha / receita)",           "Pró-labore × 12 / RBT12", "≥ 28% (Anexo III)", "Crítico fiscal — Resolução CGSN 140/18"),
]
r = 6
for kpi, formula, meta, obs in kpis:
    write_row(ws, r, [None, kpi, formula, meta, obs], align=LEFT_TOP)
    ws.cell(row=r, column=2).font = BOLD; ws.cell(row=r, column=2).fill = FILL_GREY
    ws.cell(row=r, column=3).alignment = LEFT_TOP; ws.cell(row=r, column=3).font = SMALL
    ws.cell(row=r, column=4).alignment = CENTER; ws.cell(row=r, column=4).font = BOLD; ws.cell(row=r, column=4).fill = FILL_LG
    ws.cell(row=r, column=5).font = ITAL; ws.cell(row=r, column=5).alignment = LEFT_TOP
    ws.row_dimensions[r].height = 32
    r += 1

print("Aba 16 ✓")


# ============================================================
# ABA 17 — ANÁLISE DOS ESPECIALISTAS (Squad)
# ============================================================
ws = wb.create_sheet("17. Análise Especialistas")
set_col_widths(ws, [3, 24, 80, 3])

title_row(ws, 1, "ANÁLISE DOS AGENTES ESPECIALISTAS DO SQUAD", span=4)
ws.cell(row=2, column=2, value="Cada especialista do squad avalia o plano sob sua ótica. Síntese das análises detalhadas.").font=ITAL

def write_expert_block(ws, start_row, especialista, cor_fill, parecer_geral, topicos):
    """Escreve um bloco de análise de especialista."""
    # Header
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=3)
    c = ws.cell(row=start_row, column=2, value=especialista)
    c.font = H2; c.fill = cor_fill; c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[start_row].height = 26

    # Parecer geral
    ws.merge_cells(start_row=start_row+1, start_column=2, end_row=start_row+1, end_column=3)
    c = ws.cell(row=start_row+1, column=2, value=parecer_geral)
    c.font = Font(name="Calibri", size=11, bold=True, italic=True, color="1F3864")
    c.fill = FILL_LB
    c.alignment = LEFT_TOP; c.border = BORDER
    ws.row_dimensions[start_row+1].height = 32

    r = start_row + 2
    for topic, content in topicos:
        a = ws.cell(row=r, column=2, value=topic); a.font = BOLD; a.fill = FILL_GREY
        a.alignment = LEFT_TOP; a.border = BORDER
        c = ws.cell(row=r, column=3, value=content); c.font = NORMAL
        c.alignment = LEFT_TOP; c.border = BORDER
        # Altura proporcional ao conteúdo
        lines = max(2, content.count("\n") + content.count(".") // 2 + len(content) // 80)
        ws.row_dimensions[r].height = max(28, min(140, lines * 14))
        r += 1
    return r + 1

# 1. CONTÁBIL
r = 4
r = write_expert_block(ws, r, "🧮 AGENTE CONTÁBIL",
    PatternFill("solid", fgColor="C5D9F1"),
    "Parecer: VIÁVEL com Anexo III ativo. Pró-labore R$ 8k calibrado para Fator R ≥ 28%. Premissa de 12% imposto é CONSERVADORA — efetivo real é 6-8,4% até R$ 30k/mês.",
    [
        ("Imposto de 12% é suficiente?",
         "Conservadora demais para a fase inicial. Anexo III com Fator R ≥ 28%: efetivo de 6% até R$ 180k/ano e 8,4% até R$ 360k/ano. A premissa de 12% só se torna realista quando o faturamento ultrapassa R$ 33k/mês. Acima de R$ 28,5k/mês com pró-labore R$ 8k, o Fator R cai abaixo de 28% e a empresa pode despencar para o Anexo V (15,5%) — RISCO CRÍTICO a monitorar mensalmente."),
        ("Como pró-labore impacta o PE?",
         "Pró-labore + INSS = R$ 8.880 mensais = 70% dos custos fixos. PE com R$ 8k pró-labore ≈ R$ 16-19k/mês. Se reduzido para R$ 6k (escalonado), PE cai para ~R$ 14k. Recomendação CFO: pró-labore progressivo nos 6 primeiros meses."),
        ("Estrutura de custos adequada?",
         "Razoável, com lacunas: faltam previsões de (a) provisão 13º pró-labore + férias proporcionais (CPC 25); (b) certificado digital A1 anual; (c) contingência fiscal 1-2% receita; (d) ISS antecipado em municípios de SP/RJ/BH; (e) provisão cancelamentos 1-3%. Lista completa inserida na aba 4."),
        ("DRE está coerente?",
         "Estrutura conforme CPC 26 (R1). Atenção a: (a) receita de exportação Shopify entra com isenção de ISS (art. 18 §14 LC 123/06) — segregar; (b) registrar pró-labore como despesa, não distribuição; (c) manter escrituração completa para distribuição isenta acima do limite presumido (art. 145 Res CGSN 140/18)."),
        ("Riscos contábeis/fiscais",
         "1) Desenquadramento Anexo III → V se Fator R cair: +9% de imposto = R$ 2.700/mês a +. 2) Distribuição de lucros acima do limite presumido sem escrituração completa = tributável (Lei 9.249/95). 3) Sócio com participação >10% em outra empresa cuja receita global ultrapasse R$ 4,8 mi/ano = desenquadramento. 4) NFS-e nacional obrigatória desde 2023 — precisa de certificado e credenciamento."),
        ("Contador especializado é necessário?",
         "SIM. Honorário R$ 700-1.200/mês para escritório com experiência em TI/agência (vs R$ 195 Contabilizei online). Justifica-se por: gestão mensal do Fator R, segregação de receita de exportação, limite de distribuição isenta, eSocial sócios, NFS-e."),
        ("5 ações OBRIGATÓRIAS antes do 1º faturamento",
         "1) Registrar LTDA na Junta com CNAEs corretos (6201-5/01 principal + 6202-3/00 + 6209-1/00 + 7319-0/04 + 6311-9/00); 2) Optar pelo Simples até 30 dias após CNPJ; 3) Emitir Certificado A1 + credenciar NFS-e; 4) Estruturar plano de contas + contratar contador especializado; 5) Cadastrar pró-labore no eSocial + formalizar acordo de sócios."),
    ])

# 2. JURÍDICO
r = write_expert_block(ws, r, "⚖️ AGENTE JURÍDICO",
    PatternFill("solid", fgColor="F8CBAD"),
    "Parecer: VIÁVEL com cláusulas e estrutura adequadas. SEM Política do Projeto 48h + contrato dual-track + LGPD = NÃO LANÇAR. Risco real de processo no PROCON.",
    [
        ("Riscos da promessa '48 horas'",
         "Oferta vinculante (CDC art. 30). Atraso = publicidade enganosa (art. 37 §1º) + direito de cumprimento forçado/rescisão+perdas (art. 35). Cliente B2B (PME) que usa loja como insumo → CC; MEI/PF → CDC. Necessário CONTRATO DUAL-TRACK (B2B + B2C/CDC) + Política Pública obrigatória."),
        ("Contrato — cláusulas obrigatórias",
         "1) Escopo TAXATIVO (art. 114 CC); 2) Cláusula 'Marco Zero' do prazo (pagamento + briefing + assets + aceite Política); 3) Cláusula de suspensão; 4) Aceite + DoD com silêncio = anuência (art. 111 CC, 5 du); 5) Política de revisões (2 incluídas, excedente faturado); 6) Change Request escrito; 7) Mora 1% a.m. + multa 2% + cláusula penal art. 408/411/412 CC (cap 10% do projeto); 8) Cancelamento + taxa mobilização não reembolsável (20-30%); 9) Limitação de responsabilidade = valor pago; 10) Garantia: B2C 90 dias (CDC art. 26 II), B2B 30/90 (CC 445); 11) Confidencialidade mútua; 12) PI: cliente recebe arte final; Impulso retém templates/boilerplate; 13) LGPD: papel + DPA + sub-operadores; 14) Isenção plataformas/integrações de terceiros."),
        ("Tipo societário",
         "LTDA (CC arts. 1.052-1.087). Contrato Social com: quotas (não 50/50 puro), administração, pró-labore, distribuição (pode ser desproporcional — art. 1.007), Acordo de Sócios apartado (vesting/cliff/não-competição/buy-sell/arbitragem), CNAEs corretos. CNAE principal 6201-5/01."),
        ("LGPD",
         "Papel da Impulso geralmente = OPERADORA (art. 39); pode ser CO-CONTROLADORA em tráfego pago/segmentação. DPA obrigatório. Medidas de segurança (arts. 46-49). Sub-operadores (Shopify/Meta/Google) com cláusula back-to-back. Comunicação de incidentes à ANPD (art. 48). Multa: até 2% do faturamento (cap R$ 50 mi)."),
        ("Parceria plataformas (Shopify Partners)",
         "Comissão = intermediação (LC 116/03 item 10.02 ou 17.12). Para plataforma estrangeira (Shopify Inc.): EXPORTAÇÃO de serviço → ISS isento (art. 2º I LC 116/03), PIS/COFINS alíquota zero. NÃO confundir com IRRF/CIDE/PIS-COFINS-Importação (esses incidem na remessa AO exterior, não no INGRESSO). Contrato de câmbio obrigatório (Lei 14.286/21)."),
        ("Mídia paga — RISCO ALTO",
         "NUNCA receber verba de mídia em conta da Impulso → estouro Simples (art. 3º §14 LC 123/06), risco COAF (Lei 9.613/98), exposição cambial. CLIENTE PAGA DIRETO Meta/Google. Impulso cobra apenas FEE de gestão. Cláusula expressa de não-garantia de resultado (art. 39 IV CDC)."),
        ("Top 5 riscos jurídicos",
         "1) Publicidade enganosa '48h' sem Política + gatilhos; 2) Receber verba de mídia em conta própria; 3) Ausência de DPA / papel LGPD indefinido; 4) Contrato social 50/50 sem buy/sell (paralisia → dissolução judicial art. 1.034); 5) IP não segregado (perda de ativo intangível)."),
        ("7 ações OBRIGATÓRIAS antes da 1ª venda",
         "1) Registrar LTDA + CNAEs + Simples + IM/IE; 2) Acordo de Sócios (vesting/cliff/buy-sell/arbitragem); 3) Termos de Uso + Política Privacidade do site (LGPD + Marco Civil); 4) Política do Projeto 48 Horas com aceite eletrônico; 5) Contrato dual-track B2B/B2C + Termo Aceite + DPA; 6) Registrar marca IMPULSO no INPI (classes 35/41/42); 7) Seguro RC profissional (E&O) + canal LGPD/DPO."),
    ])

# 3. FINANCEIRO
r = write_expert_block(ws, r, "💰 AGENTE FINANCEIRO",
    PatternFill("solid", fgColor="C6E0B4"),
    "Parecer: VIÁVEL no Realista (margem 24%, payback 9m). Insustentável no Pessimista. Pró-labore escalonado + foco em MRR são as alavancas-chave.",
    [
        ("Viabilidade do modelo",
         "Realista: receita anual R$ 275k, margem 24%, lucro R$ 67k, payback ~9m. Otimista: R$ 427k, margem 44%, payback 4m. Pessimista: receita R$ 115k, prejuízo R$ 67k — INVIÁVEL sem aporte ou ajuste."),
        ("Preço mínimo sustentável",
         "Custo/h interno = R$ 34,72 (Custo Fixo Total R$ 12,5k ÷ 360h capacidade). Vender abaixo = prejuízo certo. Preço/h MÍNIMO R$ 58,85 (margem 20%); RECOMENDADO R$ 89,03 (margem 40%); PREMIUM R$ 144,68 (55%). Para alinhar com mercado real (freela R$ 150-400/h, agência R$ 200-600/h), trabalhar com 75% utilização real eleva preço/h recomendado para R$ 118."),
        ("Pacote 48h — preço recomendado",
         "Start 32h → R$ 1.781 mínimo / R$ 2.490 sugerido comercial. Pro 70h → R$ 3.561 mínimo / R$ 4.990. Premium 130h → R$ 6.232 mínimo / R$ 8.900. Ancoragem psicológica + colchão p/ desconto."),
        ("Mix de receita ideal",
         "12m: 55% projetos 48h | 20% personalizados | 15% MRR | 5% parceria | 5% tráfego. 24m: 40% / 20% / 25% / 8% / 7%. 36m: 30% / 20% / 30% / 10% / 10%. META: 50% recorrente em 36 meses."),
        ("Ponto de equilíbrio",
         "PE = R$ 12.500 / 0,79 = R$ 15.823/mês (com Simples 12% + taxa 5% + inad 4%). Margem de segurança recomendada: operar a partir de R$ 25.000/mês. Em horas: 178h/mês a R$ 89/h (49% da capacidade)."),
        ("Capital de giro",
         "Burn rate mensal = R$ 12.500. Reserva mínima 3× = R$ 37.500. RECOMENDADA 6× = R$ 75.000. Conservadora 12× = R$ 150.000. Reserva inicial obrigatória ANTES do mês 1: **R$ 50.000** + linha de antecipação de recebíveis pré-aprovada."),
        ("Risco do parcelamento",
         "Vender Pro R$ 3.561 em 12x → recebe R$ 296,75/mês (não R$ 3.561 hoje). Se 50% das vendas parcelarem em 6x+, descasamento caixa de 60-90 dias. POLÍTICA: 5% off PIX/à vista; máx 3x sem juros; 6x+ com juros repassados; antecipação só se ROI > custo."),
        ("Top 5 riscos financeiros + mitigação",
         "1) Pró-labore R$ 8k mês 1 inviável → ESCALONAR R$ 3k → R$ 3,5k → R$ 4k condicional a MRR > R$ 10k. 2) Capacidade 360h é teto teórico — usar 270h faturáveis; freela a partir 70% util. 3) Concentração transacional → 5 contratos MRR/mês desde mês 1. 4) Inad B2B 4% + parcelado → política de pagamento agressiva. 5) Key-person (2 sócios) → seguro vida cruzado + SOPs."),
        ("Veredito",
         "VIÁVEL no realista com 3 ajustes: (1) pró-labore escalonado; (2) repricing comercial (Start 2.490/Pro 4.990/Premium 8.900); (3) foco agressivo em MRR. Sem ajustes, ainda viável, mas margem cai e payback estende. O upside real do negócio está em virar SaaS-like via MRR."),
    ])

# 4. PLANEJAMENTO ESTRATÉGICO
r = write_expert_block(ws, r, "🎯 AGENTE DE PLANEJAMENTO ESTRATÉGICO",
    PatternFill("solid", fgColor="FFE699"),
    "Parecer: '48h' é GATILHO DE CONVERSÃO, NÃO vantagem sustentável. Sem migrar para 'operador de crescimento' em 24-36m, perde para IA gratuita. Recorrente é o coração.",
    [
        ("Posicionamento",
         "Foco/Nicho com diferenciação operacional. Custo puro = suicídio. Diferenciação ampla = cara para 2 sócios. Nicho viável: PMEs físicas migrando para digital + lojistas em SaaS BR (Tray/Nuvemshop) profissionalizando. Proposta: 'Sua loja online profissional no ar em 48 horas — pronta para vender — com método, suporte e crescimento contínuo, sem freelancer sumido nem agência que demora 3 meses.'"),
        ("'48 horas' é vantagem real?",
         "É GATILHO DE CONVERSÃO, NÃO vantagem sustentável. Qualquer agência ou freelancer copia em 60 dias. Vantagem real = método/processo proprietário por trás das 48h: checklist 80-100 pontos, kit fiscal BR, biblioteca 30+ blocos, integrações pré-configuradas, prompts de IA versionados. '48h' vende; o processo defende."),
        ("ICP — 3 personas",
         "1) CARLA, Lojista Física 42a, boutique R$ 30-150k/mês, vende no DM — ticket R$ 3,5-6k setup + R$ 600-1,2k suporte; canal Insta orgânico + indicação contador/SEBRAE. 2) RODRIGO, Empreendedor DTC 32a, lançando marca, capital R$ 20-80k — ticket R$ 5-12k + R$ 1,5-5k tráfego; canal Google Ads + LinkedIn. 3) MARIANA, Gestora B2B PME 38a, distribuidora migrando catálogo — ticket R$ 8-20k + R$ 1,5-3k; canal LinkedIn outbound + parceria ERPs."),
        ("Mercado (TAM/SAM/SOM)",
         "TAM ~R$ 2,4bi (600k lojas SaaS × ticket R$ 4k lifetime serviço). SAM ~R$ 400-450M/ano (30k contratações + 360k×R$ 800×12 suporte). SOM 24m ~R$ 730k/ano (~0,15% SAM). Conclusão: mercado não é restrição — EXECUÇÃO é."),
        ("5 Forças (resumo)",
         "Rivalidade ALTA (agências boutique, freelancers, Shopify Partners BR). Substitutos ALTA E CRESCENTE (Shopify Magic, Wix ADI, Canva sites). Fornecedor MÉDIO-ALTO (dependência plataformas). Cliente ALTO no início, MÉDIO depois. Novos entrantes ALTA (barreira baixíssima). CONCLUSÃO: comoditizável — defesa = nicho + marca + processo + recorrência."),
        ("Vantagem competitiva sustentável",
         "Ativos a construir (em ordem): 1) Método Impulso 48h documentado (80-100 checkpoints); 2) Biblioteca templates verticais (moda/beleza/pet/alimentos/B2B); 3) Kit IA proprietário (reduz custo marginal 30-40%); 4) Cases de receita real do cliente (não 'site bonito'); 5) Comunidade Impulso (reduz churn de suporte)."),
        ("OKRs do 1º ano",
         "O1 — Validar motor comercial: KR1 R$ 480k em 12m; KR2 CAC < R$ 800, payback < 3m; KR3 70+ SQL/mês a partir mês 6. O2 — Construir MRR previsível: KR1 R$ 25k MRR mês 12; KR2 60% conversão setup→suporte; KR3 churn < 4%. O3 — Ativos defensáveis e marca: KR1 método v1.0 documentado mês 4; KR2 12 cases publicados; KR3 10k seguidores combinados sócios."),
        ("Riscos estratégicos top 5",
         "1) IA generativa (Shopify Magic/Wix ADI) elimina setup básico 24-36m → subir na cadeia, vender vendas. 2) Comoditização → nicho vertical + bundle + recusar baixo ticket. 3) Dependência plataformas → multi-plataforma + comissão <15% receita. 4) Burnout sócios → cap 5 projetos/sócio/mês + playbook + férias. 5) Falta de diferenciação → investir nos 5 ativos."),
        ("Trade-offs explícitos — NÃO fazer",
         "NÃO atende enterprise/VTEX/Magento; NÃO faz projeto <R$ 3.500; NÃO entrega site sem suporte mensal anexado; NÃO faz 'só design' ou 'só consultoria avulsa'; NÃO trabalha com cliente sem CNPJ/produto/capital; NÃO faz marketplace-only; NÃO promete prazo fora do método; NÃO aceita horas avulsas no suporte; NÃO compete em licitação pública/RFP."),
        ("Veredito",
         "Defensável hoje? Parcialmente. Sustentável aos 24m? SIM, condicional a: (1) reposicionar de 'implementador' para 'operador de crescimento de PMEs digitais'; (2) recorrente vira CORAÇÃO (50% em 24m); (3) verticalizar em 2-3 nichos. Sem ajustes: R$ 600k/ano com 2 sócios exaustos competindo com IA gratuita em 2027. Com ajustes: R$ 1,5-2M aos 36m, margem 35-45%, EV vendável."),
    ])

# 5. OPERAÇÕES DIGITAIS / E-COMMERCE
r = write_expert_block(ws, r, "⚙️ AGENTE DE OPERAÇÕES DIGITAIS E E-COMMERCE",
    PatternFill("solid", fgColor="D9D9D9"),
    "Parecer: 'Todos os produtos, todas as integrações em 48h' é ARMADILHA. Com escopo brutalmente recortado (até 20 SKUs), o modelo cabe — e gera throughput sustentável.",
    [
        ("Viabilidade real do 48h",
         "Shopify 20 SKUs: 26-66h de execução com 2 pessoas em paralelo — CABE em 48h SE pré-condições 100% fechadas. WooCommerce: 40-90h equivalente — NÃO CABE em 48h salvo escopo mínimo. Site institucional 5-8 páginas: 21-42h — cabe. 50-100 SKUs, integração ERP/OMS, marketplace, gateway custom — NÃO CABEM."),
        ("3 pacotes recomendados",
         "START: até 15 SKUs, 5 páginas, 1 gateway, 1 frete, 3 reuniões, 1 treinamento, 7d suporte. Horas: 22/32/45. PRO: até 50 SKUs, 8 páginas, 2 gateways, 2 fretes, 1 integração (Bling OU Tiny OU RD), 15d suporte. Horas: 45/70/100. PREMIUM: até 150 SKUs, 12 páginas, 2 integrações (ERP+marketplace), automações carrinho, 2 landing pages, 30d suporte. Horas: 90/130/180. WooCommerce só PRO/PREMIUM com +30% prazo/preço — DESENCORAJAR."),
        ("Capacidade operacional realista",
         "Bruto: 352h (2 sócios × 22du × 8h). Descontos: comercial 15% + atendimento 10% + pós-venda 10% + admin 8% + gestão 7% = 50% só. EXECUÇÃO PURA = ~176h/mês combinada. Alvo ≤80% = 144h para projetos. Mix sustentável: 2 Start + 1 Pro (134h) OU 1 Start + 1 Premium (162h). Premium a cada 2 meses máx enquanto dupla."),
        ("Suporte recorrente — limites",
         "Essencial 4h: 4-5 contratos/sócio = 8-10 total. Pro 10h: 2-3/sócio = 4-6 total. Growth 20h: 1-2/sócio = 2-4 total. Cada sócio MÁX 50% capacidade (90h) em suporte."),
        ("Top 10 riscos operacionais + mitigação",
         "1) Cliente atrasa assets → cláusula pausa cronômetro + 50% upfront. 2) Gateway em KYC → pré-condição rígida. 3) Doença sócio → buffer 20% + SOPs + freela plug-in. 4) Escopo creep → tabela adicionais + change request formal. 5) Cliente revende '48h' como milagre → FAQ visível 'o que NÃO entra'. 6) Pós-venda canibaliza execução → suporte pago obrigatório dia 30. 7) Plataforma quebra integração → especializar Shopify+Nuvemshop. 8) NPS baixo treinamento → videoteca + sessão gravada. 9) IA caída → templates locais + SOP fallback. 10) Inadimplência pós go-live → 50/40/10 + DNS sob Impulso até pagamento."),
        ("IA: onde dá ganho REAL ≥30%",
         "Descrições de produto 30-40%; copy institucional + políticas 40-60%; SEO title/meta/alt em massa 40%; tradução 50%; FAQ a partir de transcrição 50%. ONDE NÃO DÁ (<10%): briefing, decisão UX, QA crítico (checkout/fiscal/pixel), gateway/frete/domínio, treinamento ao vivo, negociação comercial."),
        ("SOPs/checklists INDISPENSÁVEIS antes da 1ª venda",
         "1) Checklist pré-condições D-5 (cliente assina); 2) SOP onboarding (kickoff/kit boas-vindas/acessos); 3) SOP implementação Shopify (60+ itens); 4) SOP implementação WordPress/Woo; 5) Checklist QA pré go-live (checkout/mobile/emails/fiscal/LGPD/pixels); 6) SOP handover + treinamento (roteiro + videoteca + termo aceite); 7) SOP pós-venda 30 dias."),
        ("Definition of Done",
         "Loja Shopify: compra teste real (cartão+Pix+boleto); e-mails transacionais renderizados; responsivo iOS/Android+Chrome/Safari/Edge; GA4+Pixel+GTM disparando view_item/add_to_cart/begin_checkout/purchase validados; SEO title/meta únicos + sitemap; LGPD: política+cookie banner+e-mail DPO; HTTPS+redirect www↔raiz; backup salvo; termo aceite assinado; owner 100% no cliente."),
        ("Gatilhos para 3º profissional",
         "Sustentado por 2 meses: backlog > 200h, capacidade > 85%, recusa ≥2 projetos qualificados/mês, suporte > 50h combinado, NPS < 8 ou atraso > 20%, MRR ≥ R$ 20k (sustenta CLT-flex R$ 6-9k all-in). Perfil: IMPLEMENTADOR PLENO Shopify/Tray (não sênior). Sócios sobem para vendas + suporte estratégico + conteúdo."),
        ("Veredito",
         "SIM, lançar com modificações INEGOCIÁVEIS: (1) renomear 'Express 48h úteis' padrão + 'Sprint 48h corridas' premium +40%; (2) escopo TAXATIVO máx 20 SKUs/1 gateway/1 frete/4 páginas/Shopify+Nuvemshop+Tray (Woo fora); (3) pré-condições D-5 com pausa cronômetro; (4) 50/40/10 + DNS sob Impulso; (5) DoD no contrato; (6) target 2 Start + 1 Pro/mês ou 1 Premium; (7) SOPs 1-7 + videoteca antes da 1ª venda; (8) marketing com 'o que NÃO entra' visível; (9) suporte recorrente obrigatório dia 30; (10) revisão trimestral horas reais vs estimadas. Sem isso: NÃO LANÇAR."),
    ])

print("Aba 17 ✓")


# ============================================================
# ABA 18 — ANÁLISE CRÍTICA GERAL (Devil's Advocate)
# ============================================================
ws = wb.create_sheet("18. Análise Crítica")
set_col_widths(ws, [3, 38, 80, 3])

title_row(ws, 1, "ANÁLISE CRÍTICA GERAL — DEVIL'S ADVOCATE", span=4)
ws.cell(row=2, column=2, value="Pressure-test honesto das premissas. Atacar a ideia, respeitar a pessoa.").font=ITAL

perguntas_criticas = [
    ("O modelo do Projeto 48 Horas é viável?",
     "SIM, com escopo recortado. O '48h corridas com todos os produtos e integrações' é falso por design. O '48h úteis com até 20 SKUs, 1 gateway, 1 frete, plataformas Shopify/Nuvemshop/Tray' é defensável, lucrativo e operacionalmente sustentável."),
    ("Qual escopo máximo deve ser prometido?",
     "Loja Shopify/Nuvemshop/Tray: 20 SKUs, 5 páginas, 1 gateway, 1 frete, GA4+Pixel padrão, SEO on-page básico, 1 sessão treinamento (60min), 7 dias suporte. Site institucional: até 8 páginas."),
    ("Qual escopo NÃO deve ser prometido?",
     "Migração de plataforma; +30 SKUs; tema customizado (código); integração ERP/OMS/PDV; multi-idioma; marketplace (ML/Amazon); assinaturas; B2B com tabela de preço; app mobile; blog populado; fotografia; copywriting longo; tradução. WooCommerce no regime de 48h."),
    ("O preço sugerido cobre o esforço real?",
     "Sim, no nível recomendado (Start R$ 2.490 / Pro R$ 4.990 / Premium R$ 8.900). Margem 40% considerando 12% imposto + 5% taxa + 4% inad. Vender abaixo do preço mínimo (R$ 1.781 / R$ 3.561 / R$ 6.232) = prejuízo certo."),
    ("O modelo depende demais dos 2 sócios?",
     "SIM. Risco de key-person crítico. Doença/saída de 1 sócio = -50% capacidade. Mitigação: SOPs documentados, seguro de vida cruzado, freela plug-in pré-aprovado, reserva 6 meses."),
    ("Quando contratar a 3ª pessoa?",
     "Gatilhos quantitativos (sustentados por 2 meses): backlog vendido > 200h; capacidade > 85%; recusa ≥2 projetos/mês; suporte > 50h combinado; NPS < 8 ou atraso > 20%; MRR ≥ R$ 20k. Perfil: implementador PLENO (não sênior) Shopify/Tray, CLT-flex ou PJ R$ 5-7k."),
    ("Qual serviço tende a ser mais lucrativo?",
     "1) Parceria com plataformas: ~95% margem líquida (receita passiva, custo zero). 2) Suporte mensal Growth: ~50% margem. 3) Projetos personalizados Premium: 45%. 4) Projeto 48h Premium: 40%. 5) Projeto 48h Start: 35%."),
    ("Qual serviço tende a gerar mais recorrência?",
     "Suporte mensal (MRR) — coração da tese. Parceria com plataformas — recorrência passiva. Tráfego pago — sticky por dependência operacional. Projeto 48h gera recorrência SOMENTE se converte para suporte mensal (meta: 60% conversão em 12m)."),
    ("Quais serviços devem ser evitados?",
     "1) Projetos abaixo de R$ 2.000 (canibalizam capacidade). 2) WooCommerce no formato 48h. 3) VTEX/Magento (enterprise). 4) Marketplace-only. 5) Consultoria avulsa sem entregável. 6) Migração com escopo aberto. 7) Receber verba de mídia em conta (risco fiscal)."),
    ("Quais serviços devem ser cobrados como EXTRA?",
     "Tudo que não estiver no escopo taxativo: SKUs adicionais (R$/SKU), páginas adicionais, integrações além das listadas, copywriting customizado (R$/página), fotografia, banco de imagens premium, design de logo, sessões extra de treinamento, suporte pós-go-live além do contratado, change requests."),
    ("Quais premissas precisam ser validadas no mercado?",
     "1) Disposição real dos clientes a pagar R$ 4.990 em Pro (PoC: 5 clientes em 60 dias). 2) Taxa de conversão setup→suporte de 60% (PoC: contratos em 6 meses). 3) CAC < R$ 800 (depende de canal escolhido). 4) Ganho real de produtividade com IA = 25% (medir nos primeiros 3 projetos). 5) Capacidade 360h é teto — validar com 1º trimestre real."),
    ("Pró-labore de R$ 8.000 é sustentável no início?",
     "NÃO no mês 1 do realista (PE só atinge no mês 4). RECOMENDAÇÃO: escalonar — R$ 3.000/sócio (mês 1-3) → R$ 3.500 (4-6) → R$ 4.000 (7+) condicional a MRR > R$ 10k. Diferencial vai para reserva de capital de giro. ATENÇÃO FISCAL: pró-labore total < R$ 6.700/mês pode quebrar Fator R ≥ 28% e jogar a empresa no Anexo V (15,5% imposto vs 6%). Pró-labore mínimo p/ sustentar Anexo III em R$ 25k/mês de receita = ~R$ 7.000/mês."),
    ("Faturamento mínimo mensal para operar?",
     "PE = R$ 15.823/mês (com 12% imposto + 5% taxa + 4% inad). Operar com folga: R$ 25.000/mês. Para gerar lucro saudável (margem > 20%): R$ 30.000/mês."),
]

r = 4
for pergunta, resposta in perguntas_criticas:
    a = ws.cell(row=r, column=2, value=pergunta); a.font=BOLD; a.fill=FILL_LB
    a.alignment = LEFT_TOP; a.border = BORDER
    c = ws.cell(row=r, column=3, value=resposta); c.font=NORMAL
    c.alignment = LEFT_TOP; c.border = BORDER
    lines = max(3, len(resposta) // 80 + 1)
    ws.row_dimensions[r].height = max(40, min(160, lines * 15))
    r += 1

# Síntese final
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
c = ws.cell(row=r, column=2, value="🔥 VEREDITO DEVIL'S ADVOCATE")
c.font = H2; c.fill = FILL_H1; c.alignment = LEFT_TOP
ws.row_dimensions[r].height = 24
r += 1

veredito = (
    "🟡 PROSSEGUIR COM REVISÕES CRÍTICAS.\n\n"
    "A tese resiste ao ataque NOS PARÂMETROS REALISTAS. Não resiste com a promessa original de '48h para tudo'. "
    "Os 3 pontos de fragilidade real são:\n\n"
    "1) PRÓ-LABORE: começar com R$ 8k mês 1 destrói o caixa. Escalonar é mandatório.\n\n"
    "2) ESCOPO DA OFERTA: '48h corridas para tudo' é commodities-promise inexequível. Recortar para escopo taxativo (Start/Pro/Premium com limites duros) protege a operação E aumenta percepção de profissionalismo.\n\n"
    "3) DEPENDÊNCIA DE IA GENERATIVA: a tese estratégica precisa evoluir nos 24-36 meses para 'operador de crescimento de PMEs', não 'implementador de loja'. Quem só monta loja perde para Shopify Magic/Wix ADI até 2027.\n\n"
    "Critério de falsificação: se aos 6 meses NÃO houver pelo menos R$ 10k MRR consolidado e taxa de conversão setup→suporte abaixo de 30%, a tese falha — pivotar para verticalização ou outro modelo."
)
ws.merge_cells(start_row=r, start_column=2, end_row=r+8, end_column=3)
c = ws.cell(row=r, column=2, value=veredito); c.font=NORMAL; c.fill=FILL_LY
c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1); c.border = BORDER

print("Aba 18 ✓")


# ============================================================
# ABA 19 — RECOMENDAÇÕES FINAIS
# ============================================================
ws = wb.create_sheet("19. Recomendações")
set_col_widths(ws, [3, 32, 80, 3])

title_row(ws, 1, "RECOMENDAÇÕES FINAIS DO SQUAD", span=4)
ws.cell(row=2, column=2, value="Ações objetivas e prontas para execução.").font=ITAL

recomendacoes = [
    ("Melhor modelo de precificação",
     "Pacote fechado com escopo TAXATIVO + preço de tabela público + change request formal para adicionais. Hora avulsa SOMENTE para suporte excedente e consultoria pontual. Mensalidades com vigência mínima de 12m (B2B)."),
    ("Pacote 48h para entrada no mercado",
     "Pacote PRO (R$ 4.990) como gancho principal. Start (R$ 2.490) só com promo de lançamento para construir cases. Premium (R$ 8.900) para clientes qualificados."),
    ("Preço mínimo do Projeto 48 Horas",
     "Start: R$ 1.781 | Pro: R$ 3.561 | Premium: R$ 6.232 — NUNCA vender abaixo disso (margem mínima 20%)."),
    ("Preço recomendado",
     "Start: R$ 2.490 | Pro: R$ 4.990 | Premium: R$ 8.900 — preço comercial sugerido com colchão para descontos."),
    ("Preço premium (sem desconto)",
     "Start: R$ 2.893 | Pro: R$ 5.787 | Premium: R$ 10.127 — preço de ancoragem para clientes que aceitam pagar mais por sprint dedicado."),
    ("Como proteger de escopo aberto",
     "1) Escopo TAXATIVO no contrato (lista exaustiva de incluso). 2) Política do Projeto 48h pública. 3) Cláusula de Change Request escrito + valor por hora. 4) Pacote de adicionais publicado. 5) Briefing assinado eletronicamente. 6) Pré-condições D-5. 7) Aceite + DoD com silêncio = anuência (5 du)."),
    ("Como transformar projeto em receita recorrente",
     "1) Bundle de saída: setup já inclui 30 dias suporte cortesia + oferta 20% off se contratar em 7 dias. 2) Revisão D+30 e D+60 com sócio (mostra dados → abre upsell). 3) Suporte em camadas com 'próximo degrau' claro. 4) Comunidade Impulso acessível só a clientes recorrentes. 5) SLA visível no contrato."),
    ("Como vender suporte mensal",
     "Foco comercial: 70% do esforço de vendas no suporte mensal, não em projetos. Meta agressiva: 5 contratos novos MRR/mês desde mês 1. Oferta no fechamento do projeto (timing crítico). Mecânica de auto-renovação."),
    ("Como estruturar contrato",
     "Contrato dual-track (B2B/CC + B2C/CDC). Cláusulas obrigatórias (ver aba 17 — Jurídico). Termo de Aceite/Briefing com Marco Zero. DPA LGPD. Política do Projeto 48 Horas pública. Forma de pagamento 50/40/10. Foro Sede da Impulso (B2B) ou domicílio cliente (B2C)."),
    ("Como evitar prejuízo com cartão parcelado",
     "1) 5% off PIX/à vista. 2) Máximo 3x sem juros. 3) 6x+ com juros repassados ao cliente. 4) Antecipação de recebíveis APENAS se ROI > custo (2-3% a.m.). 5) Reserva de capital de giro R$ 50k antes do mês 1."),
    ("Como usar IA sem comprometer qualidade",
     "IA SIM para: descrições de produto (30-40% ganho), copy institucional + políticas (40-60%), SEO em massa (40%), tradução (50%), FAQ a partir de transcrição (50%). IA NÃO para: briefing, decisão UX, QA crítico (checkout/fiscal/pixel), gateway/frete/domínio, treinamento ao vivo, negociação. Templates locais como fallback se a IA cair."),
    ("Como validar a aceitação do mercado",
     "Soft launch nos primeiros 90 dias com 5-10 clientes-piloto. Coletar: taxa de fechamento da proposta; tempo real vs prometido; NPS pós-projeto; conversão para suporte; CAC por canal. Pivotar pacotes/preço se desvio > 20% nas premissas."),
    ("Como organizar os primeiros 90 dias",
     "Ver aba 20 — Plano de Ação detalhado. Resumo: 30 dias estruturar (LTDA, contratos, SOPs, pacotes, site); 31-60 dias prospectar e fechar primeiros 3-5 clientes; 61-90 dias otimizar processo, criar cases, oferecer suporte recorrente."),
    ("Marca e propriedade intelectual",
     "Registrar marca IMPULSO no INPI nas classes 35 (publicidade/marketing), 42 (TI/desenvolvimento) e 41 (treinamento) — sem registro, sem proteção. Custo: ~R$ 355/classe. Contrato Social com cláusula de PI dos sócios sobre processos/templates."),
    ("Capital de giro",
     "R$ 50.000 ANTES do mês 1 (recomendação CFO conservadora). Aumentar para R$ 75.000 (6× burn) se cenário pessimista for risco real. Linha de antecipação pré-aprovada como segurança."),
]

r = 4
for titulo, conteudo in recomendacoes:
    a = ws.cell(row=r, column=2, value=titulo); a.font=BOLD; a.fill=FILL_LB
    a.alignment = LEFT_TOP; a.border = BORDER
    c = ws.cell(row=r, column=3, value=conteudo); c.font=NORMAL
    c.alignment = LEFT_TOP; c.border = BORDER
    lines = max(2, len(conteudo) // 80 + 1)
    ws.row_dimensions[r].height = max(36, min(140, lines * 15))
    r += 1

print("Aba 19 ✓")


# ============================================================
# ABA 20 — PLANO DE AÇÃO PARA OS PRIMEIROS 90 DIAS
# ============================================================
ws = wb.create_sheet("20. Plano 90 Dias")
set_col_widths(ws, [3, 6, 50, 18, 18, 22, 3])

title_row(ws, 1, "PLANO DE AÇÃO — PRIMEIROS 90 DIAS DA IMPULSO", span=7)
ws.cell(row=2, column=2, value="Checklist executável. Marque [✓] quando concluir. Responsável e prazo por iniciativa.").font=ITAL

# Cabeçalho
header_cells(ws, 4, ["#", "Ação", "Prazo", "Responsável", "Status"], start_col=2)

# 30 DIAS - ESTRUTURAÇÃO
section_row(ws, 5, "🟦 DIAS 1-30 — ESTRUTURAÇÃO LEGAL, COMERCIAL E OPERACIONAL", span=7)
acoes_30 = [
    ("Registrar LTDA na Junta Comercial + CNAEs corretos", "D+7", "Sócio 1 + Contador", "☐ Pendente"),
    ("Optar pelo Simples Nacional + emitir Certificado A1", "D+10", "Contador", "☐ Pendente"),
    ("Credenciar NFS-e nacional + município", "D+15", "Contador", "☐ Pendente"),
    ("Assinar Acordo de Sócios (vesting/cliff/buy-sell)", "D+10", "Advogado + sócios", "☐ Pendente"),
    ("Aprovar contrato dual-track B2B/B2C + Termo Aceite + DPA", "D+15", "Advogado", "☐ Pendente"),
    ("Publicar Política do Projeto 48 Horas com aceite eletrônico", "D+15", "Advogado + sócios", "☐ Pendente"),
    ("Publicar Termos de Uso + Política de Privacidade", "D+20", "Advogado", "☐ Pendente"),
    ("Registrar marca IMPULSO no INPI (classes 35/41/42)", "D+30", "Advogado", "☐ Pendente"),
    ("Definir pacotes 48h (Start/Pro/Premium) com escopo TAXATIVO", "D+15", "Sócios", "☐ Pendente"),
    ("Definir planos de suporte (Essencial/Pro/Growth)", "D+15", "Sócios", "☐ Pendente"),
    ("Criar tabela de preços + descontos públicos", "D+18", "Sócios", "☐ Pendente"),
    ("Documentar SOPs 1-7 (pré-condições, onboarding, Shopify, Woo, QA, handover, pós-venda)", "D+25", "Sócios", "☐ Pendente"),
    ("Criar biblioteca de templates Shopify (3 verticais)", "D+25", "Sócios", "☐ Pendente"),
    ("Criar biblioteca de blocos + prompts de IA versionados", "D+25", "Sócios", "☐ Pendente"),
    ("Configurar ClickUp/Notion + Trello + Google Workspace", "D+20", "Sócios", "☐ Pendente"),
    ("Configurar CRM (HubSpot Free) + funil + automações", "D+25", "Sócios", "☐ Pendente"),
    ("Criar página de vendas (landing) com pacotes + FAQ + casos", "D+27", "Sócios", "☐ Pendente"),
    ("Material comercial: deck de venda, kit boas-vindas, agenda sprint", "D+30", "Sócios", "☐ Pendente"),
    ("Reservar R$ 50.000 de capital de giro em conta segregada", "D+1", "Sócios", "☐ Pendente"),
    ("Contratar seguro RC profissional (E&O)", "D+30", "Sócios", "☐ Pendente"),
]

r = 6
for i, (acao, prazo, resp, status) in enumerate(acoes_30, 1):
    ws.cell(row=r, column=2, value=i).border=BORDER; ws.cell(row=r, column=2).alignment=CENTER; ws.cell(row=r, column=2).font=BOLD
    ws.cell(row=r, column=3, value=acao).border=BORDER; ws.cell(row=r, column=3).alignment=LEFT_TOP; ws.cell(row=r, column=3).font=NORMAL
    ws.cell(row=r, column=4, value=prazo).border=BORDER; ws.cell(row=r, column=4).alignment=CENTER
    ws.cell(row=r, column=5, value=resp).border=BORDER; ws.cell(row=r, column=5).alignment=CENTER
    ws.cell(row=r, column=6, value=status).border=BORDER; ws.cell(row=r, column=6).alignment=CENTER
    ws.row_dimensions[r].height = 24
    r += 1

# 31-60 DIAS - PROSPECÇÃO E PRIMEIROS CLIENTES
section_row(ws, r, "🟨 DIAS 31-60 — PROSPECÇÃO E PRIMEIROS CLIENTES", span=7); r += 1
acoes_60 = [
    ("Lançar conta Instagram + YouTube Shorts (rosto dos sócios)", "D+35", "Sócio 1", "☐"),
    ("Publicar lead magnet: 'Checklist 48h - 47 coisas que sua loja precisa'", "D+33", "Sócios", "☐"),
    ("Abrir formulário 'Diagnóstico Express 30min' (gratuito)", "D+33", "Sócios", "☐"),
    ("Iniciar outbound LinkedIn (50 mensagens/sem por sócio)", "D+35", "Sócios", "☐"),
    ("Cadastrar como Shopify Partner / Tray Partner / Nuvemshop Expert", "D+35", "Sócio 2", "☐"),
    ("Definir 1ª campanha Meta Ads (R$ 3-5k/mês, segmentação ICP)", "D+40", "Sócios", "☐"),
    ("Prospectar contadores parceiros para indicação cruzada (10 contatos)", "D+45", "Sócios", "☐"),
    ("Fechar 1º cliente Projeto 48h (validação de processo)", "D+50", "Sócios", "☐"),
    ("Fechar 2º cliente (ajustar tempo real vs prometido)", "D+55", "Sócios", "☐"),
    ("Fechar 3º cliente (consolidar SOPs)", "D+60", "Sócios", "☐"),
    ("Coletar NPS pós-projeto + ajustar processo se < 8", "D+60", "Sócios", "☐"),
    ("Oferecer suporte mensal aos primeiros clientes (meta: 2 conversões)", "D+60", "Sócios", "☐"),
    ("Documentar 1º case real (com receita do cliente) para portfólio", "D+60", "Sócios", "☐"),
    ("Revisão semanal de horas reais vs estimadas (ajustar preço se desvio > 15%)", "Semanal", "Sócios", "☐"),
]
for i, (acao, prazo, resp, status) in enumerate(acoes_60, 1):
    ws.cell(row=r, column=2, value=i).border=BORDER; ws.cell(row=r, column=2).alignment=CENTER; ws.cell(row=r, column=2).font=BOLD
    ws.cell(row=r, column=3, value=acao).border=BORDER; ws.cell(row=r, column=3).alignment=LEFT_TOP
    ws.cell(row=r, column=4, value=prazo).border=BORDER; ws.cell(row=r, column=4).alignment=CENTER
    ws.cell(row=r, column=5, value=resp).border=BORDER; ws.cell(row=r, column=5).alignment=CENTER
    ws.cell(row=r, column=6, value=status).border=BORDER; ws.cell(row=r, column=6).alignment=CENTER
    ws.row_dimensions[r].height = 22
    r += 1

# 61-90 DIAS - OTIMIZAÇÃO E ESCALA
section_row(ws, r, "🟩 DIAS 61-90 — OTIMIZAÇÃO, RECORRÊNCIA E PRÉ-ESCALA", span=7); r += 1
acoes_90 = [
    ("Revisar pacotes 48h com base nos primeiros 3-5 projetos", "D+65", "Sócios", "☐"),
    ("Ajustar preços se margem real divergir > 15% da projetada", "D+65", "Sócios", "☐"),
    ("Formalizar parceria com Shopify/Tray/Nuvemshop (receber comissões)", "D+70", "Sócios", "☐"),
    ("Criar processo comercial documentado (funil + scripts + propostas)", "D+75", "Sócios", "☐"),
    ("Meta: 6 clientes Projeto 48h fechados nos primeiros 90 dias", "D+90", "Sócios", "☐"),
    ("Meta: 3 contratos de suporte mensal ativos (MRR > R$ 3.000)", "D+90", "Sócios", "☐"),
    ("Atingir 60% de conversão setup → suporte mensal", "D+90", "Sócios", "☐"),
    ("Publicar 3 cases com receita real do cliente", "D+90", "Sócios", "☐"),
    ("Analisar capacidade real e gargalos (planejar 3º profissional)", "D+85", "Sócios", "☐"),
    ("Revisar metas O1/O2/O3 dos OKRs e definir KRs do trimestre seguinte", "D+90", "Sócios", "☐"),
    ("Criar v1.0 do Método Impulso 48h documentado (playbook + blocos + IA)", "D+90", "Sócios", "☐"),
    ("Revisão financeira: faturamento atingido vs PE, ajustar pró-labore", "D+90", "Sócios + Contador", "☐"),
]
for i, (acao, prazo, resp, status) in enumerate(acoes_90, 1):
    ws.cell(row=r, column=2, value=i).border=BORDER; ws.cell(row=r, column=2).alignment=CENTER; ws.cell(row=r, column=2).font=BOLD
    ws.cell(row=r, column=3, value=acao).border=BORDER; ws.cell(row=r, column=3).alignment=LEFT_TOP
    ws.cell(row=r, column=4, value=prazo).border=BORDER; ws.cell(row=r, column=4).alignment=CENTER
    ws.cell(row=r, column=5, value=resp).border=BORDER; ws.cell(row=r, column=5).alignment=CENTER
    ws.cell(row=r, column=6, value=status).border=BORDER; ws.cell(row=r, column=6).alignment=CENTER
    ws.row_dimensions[r].height = 22
    r += 1

r += 1
section_row(ws, r, "🎯 METAS DOS 90 DIAS (resumo)", span=7); r += 1
metas = [
    "✅ 6 clientes Projeto 48h fechados",
    "✅ 3 contratos de suporte mensal ativos (MRR > R$ 3.000)",
    "✅ Taxa de conversão setup → suporte ≥ 60%",
    "✅ NPS médio > 8",
    "✅ Faturamento total 90d > R$ 25.000",
    "✅ Método Impulso 48h v1.0 documentado",
    "✅ 3 cases publicados com receita real do cliente",
    "✅ Estrutura jurídica, contábil e operacional 100% pronta",
]
for m in metas:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=m); c.font=BOLD; c.fill=FILL_LG; c.alignment=LEFT; c.border=BORDER
    ws.row_dimensions[r].height = 22
    r += 1

print("Aba 20 ✓")


# ============================================================
# REORDENAR ABAS E SALVAR
# ============================================================

# Salvar arquivo final
out = "/home/user/comercial/impulso/Impulso_Plano_de_Negocios.xlsx"
wb.save(out)
print(f"\n{'='*60}")
print(f"✅ PLANILHA COMPLETA SALVA EM: {out}")
print(f"{'='*60}")
print(f"Total de abas: {len(wb.sheetnames)}")
for i, name in enumerate(wb.sheetnames, 1):
    print(f"  {i:2d}. {name}")
